#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SEIA Document Page Downloader
==============================
Descarga todos los documentos enlazados en una pagina de documento SEIA.

Uso:
    python download_documento_seia.py "https://seia.sea.gob.cl/documentos/documento.php?idDocumento=2141618022"

    O sin argumentos (pedira la URL interactivamente).
"""

import argparse
import concurrent.futures as cf
import hashlib
import json
import mimetypes
import os
import random
import re
import sys
import time
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Optional
from urllib.parse import urlparse, urlunparse, urljoin, parse_qsl

import shutil
import subprocess
import zipfile

import asyncio

import httpx
import requests
from bs4 import BeautifulSoup

from notebooklm import (
    AuthTokens,
    NotebookLMClient,
    AuthError,
    RPCError,
    SourceAddError,
    SourceProcessingError,
    SourceTimeoutError,
    ValidationError,
)
from notebooklm.auth import fetch_tokens
from notebooklm.types import source_status_to_str

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

if load_dotenv:
    load_dotenv()

# Intentar importar modulos opcionales para archivos comprimidos
try:
    import rarfile
    # Buscar UnRAR en ubicaciones comunes si no esta en PATH
    if not shutil.which("unrar"):
        for _unrar_path in [
            r"C:\Program Files\WinRAR\UnRAR.exe",
            r"C:\Program Files (x86)\WinRAR\UnRAR.exe",
        ]:
            if os.path.exists(_unrar_path):
                rarfile.UNRAR_TOOL = _unrar_path
                break
    RARFILE_DISPONIBLE = True
except ImportError:
    RARFILE_DISPONIBLE = False

try:
    import py7zr
    PY7ZR_DISPONIBLE = True
except ImportError:
    PY7ZR_DISPONIBLE = False

# ============================================================================
# CONFIGURACION
# ============================================================================

BASE_SEIA = "https://seia.sea.gob.cl"
REQUEST_DELAY_SEC = 1.0
MAX_RETRIES = 3
RETRY_BACKOFF_SEC = 5
CHUNK_SIZE = 8192
DEFAULT_OUTPUT = Path("./downloads")
ENABLE_DOWNLOAD = True
ENABLE_NOTEBOOK_SYNC = True

NOTEBOOK_CLIENT_TIMEOUT_SEC = 120
NOTEBOOK_UPLOAD_WAIT_TIMEOUT_SEC = int(
    os.getenv("NOTEBOOK_UPLOAD_WAIT_TIMEOUT_SEC", "600") or "600"
)


def _getenv_non_negative_int(name: str, default: int) -> int:
    raw_value = os.getenv(name, str(default)).strip()
    try:
        parsed_value = int(raw_value)
    except ValueError:
        return default
    return max(0, parsed_value)


def _getenv_optional_positive_int(name: str, default: int | None) -> int | None:
    raw_value = os.getenv(name)
    if raw_value is None or not raw_value.strip():
        return default
    try:
        parsed_value = int(raw_value.strip())
    except ValueError:
        return default
    return parsed_value if parsed_value > 0 else None


NOTEBOOK_SOURCES_PER_NOTEBOOK = _getenv_non_negative_int(
    "NOTEBOOK_SOURCES_PER_NOTEBOOK",
    300,
)
NOTEBOOK_UPLOAD_LIMIT = _getenv_optional_positive_int(
    "NOTEBOOK_UPLOAD_LIMIT",
    NOTEBOOK_SOURCES_PER_NOTEBOOK or None,
)
NOTEBOOK_UPLOAD_MAX_WORKERS = max(
    1,
    int(os.getenv("NOTEBOOK_UPLOAD_MAX_WORKERS", "1") or "1"),
)
NOTEBOOK_UPLOAD_RETRY_ATTEMPTS = max(
    1,
    int(os.getenv("NOTEBOOK_UPLOAD_RETRY_ATTEMPTS", "5") or "5"),
)
NOTEBOOK_UPLOAD_RETRY_BASE_SEC = float(
    os.getenv("NOTEBOOK_UPLOAD_RETRY_BASE_SEC", "5") or "5"
)
NOTEBOOK_UPLOAD_SUBMIT_JITTER_SEC = float(
    os.getenv("NOTEBOOK_UPLOAD_SUBMIT_JITTER_SEC", "0.6") or "0.6"
)

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

FILE_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".zip", ".rar", ".7z", ".gz", ".tar", ".bz2",
    ".kmz", ".kml", ".shp", ".geojson", ".dwg", ".dxf",
    ".tif", ".tiff", ".jpg", ".jpeg", ".png", ".bmp", ".gif",
    ".csv", ".txt", ".xml", ".json",
}

CONTENT_TYPE_EXT_MAP = {
    "application/pdf": ".pdf",
    "application/msword": ".doc",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/vnd.ms-excel": ".xls",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/zip": ".zip",
    "application/x-rar-compressed": ".rar",
    "application/x-7z-compressed": ".7z",
    "application/octet-stream": ".bin",
}

# ============================================================================
# UTILIDADES
# ============================================================================

def clean_text(text):
    """Limpia texto de espacios y caracteres especiales."""
    if not text:
        return ""
    return " ".join(str(text).split())


def safe_filename(name, maxlen=200):
    """Convierte un nombre a un nombre de archivo seguro."""
    if not name:
        return "sin_nombre"
    safe = str(name)
    forbidden = '<>:"/\\|?*'
    for char in forbidden:
        safe = safe.replace(char, '_')
    if len(safe) > maxlen:
        safe = safe[:maxlen]
    safe = safe.strip()
    if not safe:
        safe = "sin_nombre"
    return safe


def shorten_with_hash(text, maxlen):
    """Recorta texto y agrega hash corto para evitar colisiones."""
    text = (text or "").strip()
    if maxlen <= 0:
        return ""
    if len(text) <= maxlen:
        return text
    digest = hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()[:8]
    keep = max(1, maxlen - len(digest) - 1)
    return f"{text[:keep]}_{digest}"


def compact_component(name, maxlen=80):
    """
    Normaliza y compacta un componente de path:
    - elimina acentos
    - colapsa espacios/guiones bajos
    - acorta con hash para estabilidad
    """
    base = safe_filename(name or "sin_nombre", maxlen=500)
    base = unicodedata.normalize("NFKD", base)
    base = base.encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"\s+", " ", base).strip()
    base = re.sub(r"[ _]+", "_", base)
    if not base:
        base = "sin_nombre"
    return shorten_with_hash(base, maxlen=maxlen)


def compute_stem_budget(parent_dir, suffix_len=0, hard_cap=None):
    """Calcula largo maximo del stem para no exceder path largo en Windows."""
    if hard_cap is None:
        hard_cap = MAX_FILENAME_LEN
    stem_cap = max(12, hard_cap - max(0, suffix_len))
    if sys.platform != "win32":
        return stem_cap
    base_len = len(str(Path(parent_dir).resolve()))
    available = WINDOWS_PATH_SOFT_LIMIT - base_len - 1 - max(0, suffix_len)
    return max(12, min(stem_cap, available))


def force_https(url):
    """Fuerza HTTPS en URLs del SEIA (http no resuelve)."""
    if url and url.startswith("http://"):
        url = "https://" + url[7:]
    return url


def normalize_url(url):
    """Normaliza una URL para deduplicacion."""
    if not url or url.startswith("javascript:"):
        return None
    url = force_https(url.strip())
    if '#' in url:
        url = url.split('#')[0]
    parsed = urlparse(url)
    path = parsed.path.strip().replace(' ', '')
    if path.startswith('//'):
        path = '/' + path.lstrip('/')
    if '/archivos/' in path.lower():
        query = ''
    else:
        query = parsed.query.strip()
        if query:
            query = '&'.join(sorted(query.split('&')))
    return urlunparse((
        parsed.scheme.lower(),
        parsed.netloc.lower(),
        path,
        '', query, ''
    ))


def format_size(size_bytes):
    """Formatea bytes a tamano legible."""
    if size_bytes is None:
        return "? MB"
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    if size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"


def format_duration(seconds):
    """Formatea segundos a duracion legible."""
    if seconds < 60:
        return f"{seconds:.1f}s"
    minutes = int(seconds // 60)
    secs = int(seconds % 60)
    if minutes < 60:
        return f"{minutes}m {secs}s"
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours}h {mins}m {secs}s"


def console_safe(text):
    """Devuelve texto seguro para imprimir en consola con codificaciones limitadas."""
    s = str(text)
    enc = getattr(sys.stdout, "encoding", None) or "utf-8"
    try:
        s.encode(enc, errors="strict")
        return s
    except Exception:
        return s.encode(enc, errors="replace").decode(enc, errors="replace")


def emit_progress(progress_callback=None, **payload):
    """Notifica progreso sin interrumpir el pipeline si falla el callback."""
    if not progress_callback:
        return
    try:
        progress_callback(payload)
    except Exception as e:
        print(f"  Advertencia progreso: {console_safe(e)}")


def extract_extension(name, url=None):
    """Extrae la extension de un nombre de archivo o URL."""
    if url and not url.startswith("javascript:"):
        try:
            suf = (Path(urlparse(url).path or "").suffix or "").lower()
            if suf in FILE_EXTENSIONS:
                return suf
        except Exception:
            pass
    if name:
        part = name.replace("\\", "/").strip().split("/")[-1]
        if part:
            suf = Path(part).suffix.lower()
            if suf and (suf in FILE_EXTENSIONS or len(suf) <= 5):
                return suf
    return ""


def normalize_known_extension(ext):
    """Normaliza extension y valida contra extensiones conocidas."""
    if not ext:
        return ""
    ext = str(ext).strip().lower()
    if not ext:
        return ""
    if not ext.startswith("."):
        ext = "." + ext
    return ext if ext in FILE_EXTENSIONS else ""


# ============================================================================
# CAPA HTTP
# ============================================================================

def create_session():
    """Crea una sesion HTTP con headers apropiados."""
    session = requests.Session()
    session.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    })
    return session


def safe_request(session, method, url, stream=False, timeout=30, **kwargs):
    """Peticion HTTP con throttle y reintentos ante 429/503."""
    last_exception = None
    for intent in range(MAX_RETRIES):
        time.sleep(REQUEST_DELAY_SEC)
        try:
            if method == "get":
                r = session.get(url, stream=stream, timeout=timeout, **kwargs)
            else:
                r = session.head(url, timeout=timeout, **kwargs)
            if r.status_code in (429, 503):
                if intent < MAX_RETRIES - 1:
                    retry_after = r.headers.get("Retry-After")
                    wait = int(retry_after) if retry_after and str(retry_after).isdigit() else (RETRY_BACKOFF_SEC * (2 ** intent))
                    print(f"      Servidor ocupado ({r.status_code}). Reintento en {wait}s...")
                    time.sleep(wait)
                    continue
            r.raise_for_status()
            return r
        except requests.exceptions.RequestException as e:
            last_exception = e
            if intent < MAX_RETRIES - 1:
                wait = RETRY_BACKOFF_SEC * (2 ** intent)
                print(f"      Error de red. Reintento en {wait}s: {e}")
                time.sleep(wait)
            else:
                raise
    raise last_exception


def estimate_file_size(session, url, timeout=10):
    """Estima tamano de archivo via HEAD request."""
    if not url or url.startswith("javascript:"):
        return None
    try:
        response = safe_request(session, "head", url, timeout=timeout, allow_redirects=True)
        if response.status_code == 200:
            cl = response.headers.get("Content-Length")
            if cl:
                return int(cl)
    except Exception:
        pass
    return None


# ============================================================================
# PARSER DE PAGINA
# ============================================================================

def fetch_page(session, url):
    """Obtiene la pagina y retorna BeautifulSoup."""
    response = safe_request(session, "get", url, timeout=30)
    try:
        soup = BeautifulSoup(response.content, "lxml")
    except Exception:
        soup = BeautifulSoup(response.content, "html.parser")
    return soup


def extract_metadata(soup):
    """Extrae metadata del proyecto desde la pagina."""
    metadata = {
        "nombre_proyecto": "",
        "tipo_documento": "",
        "empresa": "",
        "region": "",
        "inversion": "",
    }

    # El titulo del proyecto suele estar en <h2> o en campos de metadata
    # Buscar patrones comunes en paginas SEIA
    all_text = soup.get_text(separator="\n")

    # Buscar nombre del proyecto - generalmente el segundo heading grande
    headings = soup.find_all(["h1", "h2", "h3"])
    for h in headings:
        text = clean_text(h.get_text())
        # Limpiar comillas tipograficas y normales
        text = text.strip('""\u201c\u201d\u00ab\u00bb\u2018\u2019\'')
        if text and len(text) > 3 and text.lower() not in ("estudio de impacto ambiental", "declaracion de impacto ambiental"):
            metadata["nombre_proyecto"] = text
            break

    # Buscar tipo (EIA/DIA)
    for h in headings:
        text = clean_text(h.get_text()).lower()
        if "estudio de impacto ambiental" in text:
            metadata["tipo_documento"] = "Estudio de Impacto Ambiental"
            break
        if "declaracion de impacto ambiental" in text or "declaración de impacto ambiental" in text:
            metadata["tipo_documento"] = "Declaracion de Impacto Ambiental"
            break

    # Buscar campos de metadata en tablas o texto
    lines = all_text.split("\n")
    for i, line in enumerate(lines):
        line_clean = clean_text(line).lower()
        if "titular" in line_clean or "empresa" in line_clean:
            # El valor suele estar en la siguiente linea o despues de ":"
            if ":" in line:
                val = clean_text(line.split(":", 1)[1])
                if val:
                    metadata["empresa"] = val
            elif i + 1 < len(lines):
                val = clean_text(lines[i + 1])
                if val and len(val) > 2:
                    metadata["empresa"] = val
        elif ("region" in line_clean or "región" in line_clean) and not metadata["region"]:
            if ":" in line:
                val = clean_text(line.split(":", 1)[1])
                if val:
                    metadata["region"] = val
            elif i + 1 < len(lines):
                val = clean_text(lines[i + 1])
                if val and len(val) > 2:
                    metadata["region"] = val
        elif "inversion" in line_clean or "inversión" in line_clean:
            if ":" in line:
                val = clean_text(line.split(":", 1)[1])
                if val:
                    metadata["inversion"] = val
            elif i + 1 < len(lines):
                val = clean_text(lines[i + 1])
                if val and len(val) > 2:
                    metadata["inversion"] = val

    return metadata


def _find_nearest_heading(element):
    """Encuentra el heading (h2/h3/h4) mas cercano hacia arriba en el DOM."""
    # Primero buscar entre siblings anteriores
    current = element
    while current:
        current = current.find_previous(["h1", "h2", "h3", "h4", "h5"])
        if current:
            text = clean_text(current.get_text())
            if text and len(text) > 1:
                return text
    return "General"


def _is_file_link(href):
    """Determina si un href apunta a un archivo descargable."""
    if not href:
        return False
    href_lower = href.lower().strip()
    if href_lower.startswith("javascript:") or href_lower.startswith("#") or href_lower.startswith("mailto:"):
        return False
    # Verificar si contiene /archivos/ (patron SEIA)
    if "/archivos/" in href_lower:
        return True
    # Verificar extension conocida
    try:
        path = urlparse(href).path
        ext = Path(path).suffix.lower()
        if ext in FILE_EXTENSIONS:
            return True
    except Exception:
        pass
    return False


def extract_download_links(soup, page_url):
    """
    Encuentra todos los enlaces a archivos descargables en la pagina.
    Retorna lista de DocumentInfo dicts.
    """
    documents = []
    seen_urls = set()
    index = 0

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"].strip()

        if not _is_file_link(href):
            continue

        # Construir URL absoluta (forzar HTTPS)
        if href.startswith("http"):
            full_url = force_https(href)
        elif href.startswith("/"):
            full_url = BASE_SEIA + href
        else:
            full_url = force_https(urljoin(page_url, href))

        # Deduplicar
        norm = normalize_url(full_url)
        if norm in seen_urls:
            continue
        seen_urls.add(norm)

        # Extraer nombre del documento
        link_text = clean_text(a_tag.get_text())
        if not link_text:
            # Usar nombre del archivo de la URL
            try:
                link_text = Path(urlparse(full_url).path).name
            except Exception:
                link_text = f"documento_{index + 1}"

        # Determinar seccion
        section = _find_nearest_heading(a_tag)

        # Extension
        ext = extract_extension(link_text, full_url)

        index += 1
        documents.append({
            "index": index,
            "name": link_text,
            "url": full_url,
            "section": section,
            "categoria": section,
            "texto_link": link_text,
            "url_origen": full_url,
            "file_type": ext,
            "size_bytes": None,
            "download_path": None,
            "status": "pending",
            "error": None,
        })

    return documents


# ============================================================================
# ESTIMACION DE TAMANOS
# ============================================================================

def estimate_all_sizes(session, documents):
    """Estima tamano de todos los documentos via HEAD requests."""
    total = len(documents)
    estimated_count = 0
    total_size = 0

    for i, doc in enumerate(documents):
        filename = doc["name"]
        if len(filename) > 50:
            filename = filename[:47] + "..."
        print(f"  [{i+1}/{total}] {filename} ... ", end="", flush=True)

        size = estimate_file_size(session, doc["url"])
        doc["size_bytes"] = size
        if size is not None:
            estimated_count += 1
            total_size += size
            print(format_size(size))
        else:
            print("tamano desconocido")

    return total_size, estimated_count


# ============================================================================
# MOTOR DE DESCARGA
# ============================================================================

def _resolve_download_path(doc, output_dir):
    """Genera ruta de descarga, manejando duplicados."""
    section_dir = output_dir / compact_component(doc["section"], maxlen=MAX_SECTION_DIR_LEN)
    section_dir.mkdir(parents=True, exist_ok=True)

    original_name = safe_filename(doc["name"], maxlen=500)
    name_suffix = Path(original_name).suffix.lower()
    known_name_suffix = normalize_known_extension(name_suffix)
    detected_suffix = normalize_known_extension(doc.get("file_type", ""))

    # Si el nombre trae una pseudo-extension (ej: ".1 antecedentes legales"),
    # preferir la extension detectada desde URL/Content-Type.
    suffix = detected_suffix or known_name_suffix
    has_valid_suffix_in_name = bool(known_name_suffix and known_name_suffix == name_suffix)
    stem = Path(original_name).stem if has_valid_suffix_in_name else original_name
    stem_budget = compute_stem_budget(section_dir, suffix_len=len(suffix), hard_cap=MAX_FILENAME_LEN)
    filename = compact_component(stem, maxlen=stem_budget) + suffix

    filepath = section_dir / filename

    # Manejar duplicados
    if filepath.exists():
        counter = 2
        while filepath.exists():
            extra = f"_{counter}{suffix}"
            stem_budget = compute_stem_budget(section_dir, suffix_len=len(extra), hard_cap=MAX_FILENAME_LEN)
            dedup_stem = compact_component(stem, maxlen=stem_budget)
            filepath = section_dir / f"{dedup_stem}_{counter}{suffix}"
            counter += 1

    return filepath


def download_file(session, doc, output_dir):
    """
    Descarga un archivo individual.
    Retorna (path_descargado, error_o_None).
    """
    url = doc.get("url")
    if not url:
        return (None, "Sin URL")

    try:
        filepath = _resolve_download_path(doc, output_dir)
        doc["download_path"] = str(filepath.relative_to(output_dir))

        # Skip si ya existe (reanudable)
        if filepath.exists() and filepath.stat().st_size > 0:
            doc["status"] = "done"
            doc["size_bytes"] = filepath.stat().st_size
            return (filepath, None)

        os.makedirs(_long_path(filepath.parent), exist_ok=True)

        response = safe_request(session, "get", url, stream=True, timeout=60)
        response.raise_for_status()

        # Si no tiene extension, intentar deducir de Content-Type
        if not filepath.suffix or filepath.suffix == ".bin":
            content_type = response.headers.get("Content-Type", "").split(";")[0].strip()
            ext = CONTENT_TYPE_EXT_MAP.get(content_type)
            if ext and ext != ".bin":
                filepath = filepath.with_suffix(ext)
                doc["download_path"] = str(filepath.relative_to(output_dir))

        downloaded = 0
        with open(_long_path(filepath), "wb") as f:
            for chunk in response.iter_content(chunk_size=CHUNK_SIZE):
                if chunk:
                    downloaded += len(chunk)
                    f.write(chunk)

        doc["size_bytes"] = downloaded
        doc["status"] = "done"
        return (filepath, None)

    except Exception as e:
        doc["status"] = "error"
        doc["error"] = str(e)
        return (None, str(e))


def download_all(session, documents, output_dir, progress_callback=None):
    """Descarga todos los documentos con progreso detallado."""
    total = len(documents)
    done = 0
    failed = 0
    total_bytes = 0
    start_time = time.time()
    emit_progress(
        progress_callback,
        stage="downloading",
        current=0,
        total=total,
        message="Iniciando descarga de documentos SEIA.",
    )

    for i, doc in enumerate(documents):
        filename = doc["name"]
        size_str = format_size(doc["size_bytes"]) if doc["size_bytes"] else "?"
        if len(filename) > 55:
            filename = filename[:52] + "..."

        print(f"  [{i+1}/{total}] {console_safe(filename)} ({size_str}) ... ", end="", flush=True)

        file_start = time.time()
        path, error = download_file(session, doc, output_dir)
        file_elapsed = time.time() - file_start

        if error:
            failed += 1
            print(f"ERROR ({console_safe(error)})")
        else:
            done += 1
            actual_size = doc.get("size_bytes", 0) or 0
            total_bytes += actual_size
            print(f"OK [{format_duration(file_elapsed)}]")

        emit_progress(
            progress_callback,
            stage="downloading",
            current=i + 1,
            total=total,
            message=f"Descarga {i + 1}/{total}: {filename}",
            done=done,
            failed=failed,
        )

        # Progreso parcial cada 10 archivos
        if (i + 1) % 10 == 0 and i + 1 < total:
            elapsed = time.time() - start_time
            print(f"  --- Progreso: {done}/{i+1} OK, {failed} errores | "
                  f"{format_size(total_bytes)} descargados | "
                  f"Tiempo: {format_duration(elapsed)}")

    elapsed = time.time() - start_time
    return done, failed, total_bytes, elapsed


# ============================================================================
# LISTADO Y REPORTE
# ============================================================================

def list_downloaded_files(output_dir, max_depth=2):
    """Recorre la carpeta de salida y lista archivos (hasta max_depth niveles)."""
    file_list = []
    output_dir = Path(output_dir).resolve()
    for root, dirs, files in _walk_paths(output_dir):
        # Limitar profundidad para no listar miles de archivos extraidos
        rel_root = Path(root).relative_to(output_dir)
        depth = len(rel_root.parts)
        if depth > max_depth:
            continue
        for f in sorted(files):
            if f == "manifest.json":
                continue
            filepath = Path(root) / f
            try:
                rel_path = filepath.relative_to(output_dir)
            except ValueError:
                continue
            size = _safe_stat_size(filepath)
            file_list.append((str(rel_path), size))
    return file_list


def save_manifest(documents, metadata, page_url, output_dir, elapsed):
    """Guarda manifest.json con toda la informacion."""
    id_documento = ""
    parsed = urlparse(page_url)
    if parsed.query:
        for key, value in parse_qsl(parsed.query, keep_blank_values=True):
            if key.lower() in {"iddocumento", "docid"} and value:
                id_documento = value.strip("/")
                break

    done_count = sum(1 for d in documents if d["status"] == "done")
    failed_count = sum(1 for d in documents if d["status"] == "error")
    total_size = sum(d.get("size_bytes") or 0 for d in documents if d["status"] == "done")

    manifest = {
        "url": page_url,
        "id_documento": id_documento,
        "nombre_proyecto": metadata.get("nombre_proyecto", ""),
        "tipo_documento": metadata.get("tipo_documento", ""),
        "empresa": metadata.get("empresa", ""),
        "region": metadata.get("region", ""),
        "timestamp": datetime.now().isoformat(),
        "total_documentos": len(documents),
        "total_descargados": done_count,
        "total_fallidos": failed_count,
        "total_bytes": total_size,
        "tiempo_segundos": round(elapsed, 2),
        "documentos": documents,
    }

    manifest_path = output_dir / "manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    return manifest_path


def load_manifest(manifest_path):
    """Carga manifest.json existente, si existe y es valido."""
    p = Path(manifest_path)
    if not p.exists():
        return None
    try:
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except Exception:
        return None
    return None


def build_download_trace_rows(
    output_dir,
    reference_documents=None,
    extraction_levels=None,
    return_stats=False,
):
    """
    Construye filas para trazabilidad de descarga/descompresion.
    Niveles:
    - 0: archivos iniciales antes de descompresion.
    - 1+: archivos generados por cada pasada de descompresion.
    Marca PDF de primera corrida descargada.
    Si return_stats=True, retorna (rows, stats).
    """
    output_dir = Path(output_dir).resolve()
    reference_documents = reference_documents or []
    extraction_levels = extraction_levels or {}

    total_detected_files = 0
    excluded_files = 0

    downloaded_keys = set()
    first_run_pdf_keys = set()
    metadata_refs = []

    def _doc_trace_metadata(doc, full_path):
        return {
            "categoria": doc.get("categoria") or doc.get("section") or "",
            "texto_link": doc.get("texto_link") or doc.get("name") or "",
            "url_origen": doc.get("url_origen") or doc.get("url") or "",
            "archivo_descargado_raiz": Path(str(doc.get("download_path") or full_path)).name,
            "ruta_archivo_descargado_raiz": str(doc.get("download_path") or ""),
        }

    for doc in reference_documents:
        if doc.get("status") != "done":
            continue
        rel_path = doc.get("download_path")
        if not rel_path:
            continue
        full_path = output_dir / rel_path
        key = _path_key(full_path)
        downloaded_keys.add(key)
        if Path(rel_path).suffix.lower() == ".pdf":
            first_run_pdf_keys.add(key)
        metadata_refs.append({
            "key": key,
            "path": full_path,
            "extract_dir": _extract_dir_for_archive(full_path),
            "metadata": _doc_trace_metadata(doc, full_path),
        })

    def _metadata_for_file(filepath):
        key = _path_key(filepath)
        for ref in metadata_refs:
            if ref["key"] == key:
                return ref["metadata"]
        for ref in metadata_refs:
            try:
                Path(filepath).resolve().relative_to(Path(ref["extract_dir"]).resolve())
                return ref["metadata"]
            except Exception:
                continue
        return {
            "categoria": "",
            "texto_link": "",
            "url_origen": "",
            "archivo_descargado_raiz": "",
            "ruta_archivo_descargado_raiz": "",
        }

    rows = []
    for root, _, files in _walk_paths(output_dir):
        for f in files:
            total_detected_files += 1
            if f in GENERATED_LOCAL_FILENAMES:
                excluded_files += 1
                continue
            filepath = Path(root) / f
            try:
                rel_path = filepath.relative_to(output_dir)
            except ValueError:
                rel_path = filepath

            key = _path_key(filepath)
            if key in downloaded_keys:
                level = 0
                origin = "descarga"
            else:
                mapped_level = extraction_levels.get(key)
                if mapped_level is None:
                    level = 0
                    origin = "otro"
                else:
                    level = mapped_level
                    origin = "descompresion" if level > 0 else "inicial"

            size_bytes = _safe_stat_size(filepath)
            rel_path_str = str(rel_path)
            trace_metadata = _metadata_for_file(filepath)

            rows.append({
                "ruta": rel_path_str,
                "extension": filepath.suffix.lower(),
                "peso_bytes": size_bytes,
                "peso_legible": format_size(size_bytes),
                "nivel_descarga_descompresion": level,
                "pdf_primera_corrida": "SI" if key in first_run_pdf_keys else "NO",
                "origen": origin,
                # Columnas equivalentes al Excel de listado final (CP6B).
                "nombre_archivo": filepath.name,
                "ruta_relativa": rel_path_str,
                "ruta_absoluta": _safe_resolve_str(filepath),
                "tamano_bytes": size_bytes,
                "tamano_legible": format_size(size_bytes),
                "categoria": trace_metadata["categoria"],
                "texto_link": trace_metadata["texto_link"],
                "url_origen": trace_metadata["url_origen"],
                "archivo_descargado_raiz": trace_metadata["archivo_descargado_raiz"],
                "ruta_archivo_descargado_raiz": trace_metadata["ruta_archivo_descargado_raiz"],
            })

    rows.sort(key=lambda x: (x["nivel_descarga_descompresion"], x["ruta"].lower()))
    if return_stats:
        stats = {
            "archivos_detectados": total_detected_files,
            "archivos_excluidos": excluded_files,
            "archivos_finales_trazabilidad": len(rows),
        }
        return rows, stats
    return rows


def export_download_trace_excel(rows, output_dir):
    """
    Exporta trazabilidad de descarga/descompresion a Excel.
    Retorna (excel_path, error_o_None).
    """
    output_dir = Path(output_dir)
    excel_path = output_dir / "trazabilidad_descarga_descompresion.xlsx"

    try:
        from openpyxl import Workbook
    except ImportError:
        return (None, "No se pudo generar trazabilidad Excel: openpyxl no instalado (pip install openpyxl)")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trazabilidad"

        headers = [
            "ruta",
            "extension",
            "peso_bytes",
            "peso_legible",
            "nivel_descarga_descompresion",
            "pdf_primera_corrida",
            "origen",
            "nombre_archivo",
            "ruta_relativa",
            "ruta_absoluta",
            "tamano_bytes",
            "tamano_legible",
            "categoria",
            "texto_link",
            "url_origen",
            "archivo_descargado_raiz",
            "ruta_archivo_descargado_raiz",
        ]
        ws.append(headers)

        for item in rows:
            ws.append([
                item["ruta"],
                item["extension"],
                item["peso_bytes"],
                item["peso_legible"],
                item["nivel_descarga_descompresion"],
                item["pdf_primera_corrida"],
                item["origen"],
                item.get("nombre_archivo"),
                item.get("ruta_relativa"),
                item.get("ruta_absoluta"),
                item.get("tamano_bytes"),
                item.get("tamano_legible"),
                item.get("categoria"),
                item.get("texto_link"),
                item.get("url_origen"),
                item.get("archivo_descargado_raiz"),
                item.get("ruta_archivo_descargado_raiz"),
            ])

        column_widths = {
            "A": 110,
            "B": 14,
            "C": 15,
            "D": 15,
            "E": 22,
            "F": 20,
            "G": 18,
            "H": 45,
            "I": 110,
            "J": 120,
            "K": 15,
            "L": 15,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        try:
            wb.save(excel_path)
            return (excel_path, None)
        except PermissionError:
            stamped = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback_path = output_dir / f"trazabilidad_descarga_descompresion_{stamped}.xlsx"
            wb.save(fallback_path)
            return (fallback_path, None)
    except Exception as e:
        return (None, f"No se pudo generar trazabilidad Excel: {e}")


def print_file_listing(output_dir):
    """Imprime el listado de archivos descargados organizados por carpeta."""
    file_list = list_downloaded_files(output_dir)
    if not file_list:
        print("  (ninguno)")
        return

    current_folder = None
    for rel_path, size in file_list:
        parts = Path(rel_path).parts
        folder = parts[0] if len(parts) > 1 else ""
        filename = parts[-1]

        if folder != current_folder:
            current_folder = folder
            if folder:
                print(f"    {console_safe(folder)}/")

        padding = "      " if folder else "    "
        size_str = format_size(size).rjust(12)
        print(f"{padding}{console_safe(filename)}{size_str}")


def print_summary(documents, metadata, output_dir, elapsed):
    """Imprime resumen final completo."""
    done_count = sum(1 for d in documents if d["status"] == "done")
    failed_count = sum(1 for d in documents if d["status"] == "error")
    total_size = sum(d.get("size_bytes") or 0 for d in documents if d["status"] == "done")
    skipped_count = sum(1 for d in documents if d["status"] == "pending")

    print()
    print("=" * 64)
    print("RESUMEN FINAL")
    print("=" * 64)
    print(f"  Documentos encontrados:  {len(documents)}")
    print(f"  Descargados:             {done_count}  ({format_size(total_size)})")
    if failed_count > 0:
        print(f"  Fallidos:                {failed_count}")
    if skipped_count > 0:
        print(f"  Pendientes/omitidos:     {skipped_count}")
    print(f"  Carpeta:                 {output_dir}")
    print(f"  Manifest:                {output_dir / 'manifest.json'}")
    print(f"  Tiempo total:            {format_duration(elapsed)}")

    # Listado de archivos
    print()
    print("  Archivos descargados:")
    print_file_listing(output_dir)

    # Listado de errores
    failed_docs = [d for d in documents if d["status"] == "error"]
    if failed_docs:
        print()
        print("  Fallidos:")
        for d in failed_docs:
            print(
                f"    [{d['index']}] "
                f"{console_safe(d['name'])} - {console_safe(d['error'])}"
            )

    print()


# ============================================================================
# EXTRACCION DE ARCHIVOS COMPRIMIDOS
# ============================================================================

ARCHIVE_EXTENSIONS = {".rar", ".zip", ".7z", ".kmz"}
RAR_MULTIPART_RE = re.compile(r"^(?P<prefix>.+?)\.part(?P<part>\d+)\.rar$", re.IGNORECASE)
NUMERIC_DUPLICATE_SUFFIX_RE = re.compile(r"_(\d+)$")
REPORT_EXTENSIONS = {".pdf", ".doc"}
FINAL_PDF_EXCLUDE_KEYWORDS = (
    "planitmetria",
    "planimetria",
    "hds",
    "plano",
    "planos",
    "lamina",
    "laminas",
    "apendices",
    "apendice",
    "foto",
)
WINDOWS_PATH_SOFT_LIMIT = 240
MAX_SECTION_DIR_LEN = 40
MAX_FILENAME_LEN = 100
MAX_EXTRACT_DIR_LEN = 36
MAX_EXTRACT_PASSES = 8
GENERATED_LOCAL_FILENAMES = {
    "manifest.json",
    "listado_pdf_doc.xlsx",
    "listado_final_pdf.xlsx",
    "trazabilidad_descarga_descompresion.xlsx",
}


def _long_path(p):
    """Agrega prefijo \\\\?\\ para soportar paths largos en Windows."""
    p = Path(p).resolve()
    s = str(p)
    if sys.platform == "win32" and not s.startswith("\\\\?\\"):
        s = "\\\\?\\" + s
    return s


def _strip_long_path_prefix(path_str):
    """Elimina prefijo de path largo de Windows (\\\\?\\) para presentar rutas legibles."""
    if sys.platform != "win32":
        return path_str
    if path_str.startswith("\\\\?\\UNC\\"):
        return "\\\\" + path_str[8:]
    if path_str.startswith("\\\\?\\"):
        return path_str[4:]
    return path_str


def _walk_paths(base_dir):
    """
    Itera archivos/directorios con soporte de paths largos en Windows.
    Retorna tuplas (root_path: Path, dirs: list[str], files: list[str]).
    """
    base = Path(base_dir).resolve()
    walk_root = _long_path(base) if sys.platform == "win32" else str(base)
    for root, dirs, files in os.walk(walk_root, onerror=lambda _e: None):
        yield Path(_strip_long_path_prefix(root)), dirs, files


def _safe_resolve_str(filepath):
    """Convierte Path a string absoluto de forma tolerante a rutas largas."""
    p = Path(filepath)
    try:
        return str(p.resolve())
    except Exception:
        return str(p)


def _path_key(filepath):
    """Normaliza path a clave estable en minusculas."""
    return _safe_resolve_str(filepath).lower()


def _count_files_in_dir(d):
    """Cuenta archivos (no dirs) recursivamente."""
    count = 0
    for _, _, files in _walk_paths(d):
        count += len(files)
    return count


def _find_unrar_tool():
    """Busca el ejecutable UnRAR en el sistema."""
    if RARFILE_DISPONIBLE:
        tool = getattr(rarfile, "UNRAR_TOOL", None)
        if tool and os.path.exists(tool):
            return tool
    tool = shutil.which("unrar")
    if tool:
        return tool
    for p in [r"C:\Program Files\WinRAR\UnRAR.exe",
              r"C:\Program Files (x86)\WinRAR\UnRAR.exe"]:
        if os.path.exists(p):
            return p
    return None


def _robocopy_move(src, dst):
    """Mueve directorio usando robocopy (soporta paths largos en Windows)."""
    dst_str = str(Path(dst).resolve())
    src_str = str(Path(src).resolve())
    # robocopy /MOVE /E: mover todo recursivamente
    result = subprocess.run(
        ["robocopy", src_str, dst_str, "/MOVE", "/E", "/NFL", "/NDL", "/NJH", "/NJS", "/NC", "/NS", "/NP"],
        capture_output=True, text=True, timeout=300
    )
    # robocopy returns 0-7 on success, >=8 on error
    if result.returncode >= 8:
        raise OSError(f"robocopy error (code {result.returncode}): {result.stderr or result.stdout}")


def _link_or_copy(src, dst):
    """Crea hardlink si es posible, sino copia el archivo."""
    dst.parent.mkdir(parents=True, exist_ok=True)
    try:
        os.link(str(src), str(dst))
    except OSError:
        shutil.copy2(str(src), str(dst))


def _extract_rar_group_with_unrar(volume_items, start_source_name, extract_dir):
    """
    Extrae un grupo de volumenes RAR usando sus nombres originales.
    Esto permite manejar .part1/.part2 aunque los archivos locales tengan otro nombre.
    """
    import tempfile

    unrar_tool = _find_unrar_tool()
    if not unrar_tool:
        return (0, "UnRAR no encontrado. Instala WinRAR o agrega unrar al PATH.")

    stage_dir = Path(tempfile.mkdtemp(prefix="seia_vol_"))
    tmp_extract_dir = Path(tempfile.mkdtemp(prefix="seia_ext_"))

    try:
        for item in volume_items:
            src = Path(item["path"]).resolve()
            dst = stage_dir / item["source_name"]
            _link_or_copy(src, dst)

        start_archive = stage_dir / start_source_name
        if not start_archive.exists():
            return (0, f"Volumen inicial no encontrado: {start_source_name}")

        result = subprocess.run(
            [unrar_tool, "x", "-o+", "-y", str(start_archive), str(tmp_extract_dir) + os.sep],
            capture_output=True, text=True, timeout=900
        )

        num_files = _count_files_in_dir(tmp_extract_dir)
        if result.returncode != 0 and num_files == 0:
            err = result.stderr.strip() or result.stdout.strip()
            if len(err) > 300:
                err = err[:300] + "..."
            return (0, f"UnRAR error (code {result.returncode}): {err}")

        if Path(extract_dir).exists():
            shutil.rmtree(str(extract_dir), ignore_errors=True)
        os.makedirs(str(extract_dir), exist_ok=True)

        if sys.platform == "win32":
            _robocopy_move(tmp_extract_dir, extract_dir)
        else:
            shutil.move(str(tmp_extract_dir), str(extract_dir))

        num_files = _count_files_in_dir(extract_dir)
        return (num_files, None)

    except Exception as e:
        return (0, str(e))
    finally:
        if stage_dir.exists():
            shutil.rmtree(stage_dir, ignore_errors=True)
        if tmp_extract_dir.exists():
            shutil.rmtree(tmp_extract_dir, ignore_errors=True)


def _extract_rar_with_unrar(archive_path, extract_dir):
    """
    Extrae RAR usando UnRAR.exe via directorio temporal corto.
    Evita el limite de 260 caracteres de Windows extrayendo a un path corto
    y usando robocopy para mover al destino final.
    """
    item = {
        "path": Path(archive_path),
        "source_name": Path(archive_path).name,
    }
    return _extract_rar_group_with_unrar([item], item["source_name"], extract_dir)


def _discover_archive_entries(output_dir, documents=None):
    """
    Lista archivos comprimidos candidatos.
    Si se entrega `documents`, usa solo los documentos descargados en esta ejecucion.
    """
    output_dir = Path(output_dir)
    entries = []

    if documents:
        for doc in documents:
            if doc.get("status") != "done":
                continue
            rel_path = doc.get("download_path")
            if not rel_path:
                continue
            archive_path = output_dir / rel_path
            ext = archive_path.suffix.lower()
            if ext not in ARCHIVE_EXTENSIONS:
                continue
            if not archive_path.exists():
                continue
            source_name = Path(urlparse(doc.get("url", "")).path).name or archive_path.name
            entries.append({
                "path": archive_path,
                "ext": ext,
                "source_name": source_name,
                "url": doc.get("url", ""),
            })
        return entries

    for root, _, files in _walk_paths(output_dir):
        for f in files:
            if f == "manifest.json":
                continue
            archive_path = Path(root) / f
            ext = archive_path.suffix.lower()
            if ext in ARCHIVE_EXTENSIONS:
                entries.append({
                    "path": archive_path,
                    "ext": ext,
                    "source_name": archive_path.name,
                    "url": "",
                })

    return entries


def _choose_best_volume_candidate(items):
    """Elige mejor candidato cuando hay duplicados del mismo volumen (ej: _2, _3)."""
    if len(items) == 1:
        return items[0]

    def _rank(item):
        stem = Path(item["path"]).stem
        has_duplicate_suffix = 1 if NUMERIC_DUPLICATE_SUFFIX_RE.search(stem) else 0
        return (has_duplicate_suffix, len(item["path"].name))

    return sorted(items, key=_rank)[0]


def _build_extraction_plan(entries, output_dir):
    """
    Construye plan de extraccion:
    - Archivos simples (zip/7z/kmz/rar simple)
    - Grupos RAR multivolumen (.partN.rar) extraidos una sola vez desde volumen inicial.
    """
    tasks = []
    multipart_groups = {}

    for entry in entries:
        if entry["ext"] != ".rar":
            tasks.append({"kind": "single", "entry": entry})
            continue

        source_name = entry.get("source_name") or entry["path"].name
        match = RAR_MULTIPART_RE.match(source_name)
        if not match:
            tasks.append({"kind": "single", "entry": entry})
            continue

        part_num = int(match.group("part"))
        group_key = (
            str(entry["path"].parent).lower(),
            match.group("prefix").lower(),
        )
        group = multipart_groups.setdefault(group_key, {
            "prefix": match.group("prefix"),
            "by_part": {},
        })
        group["by_part"].setdefault(part_num, []).append(entry)

    for group in multipart_groups.values():
        selected_volumes = []
        for part_num in sorted(group["by_part"].keys()):
            selected_volumes.append(_choose_best_volume_candidate(group["by_part"][part_num]))

        has_part1 = 1 in group["by_part"]
        if has_part1:
            start_item = _choose_best_volume_candidate(group["by_part"][1])
        else:
            first_part = min(group["by_part"].keys())
            start_item = _choose_best_volume_candidate(group["by_part"][first_part])

        extract_folder_name = compact_component(group["prefix"], maxlen=MAX_EXTRACT_DIR_LEN)
        extract_dir = start_item["path"].parent / extract_folder_name

        tasks.append({
            "kind": "rar_multivol",
            "start_item": start_item,
            "volumes": selected_volumes,
            "has_part1": has_part1,
            "extract_dir": extract_dir,
        })

    def _task_key(task):
        if task["kind"] == "single":
            p = task["entry"]["path"]
        else:
            p = task["start_item"]["path"]
        return str(p).lower()

    tasks.sort(key=_task_key)
    return tasks


def extract_archive(archive_path):
    """
    Extrae un archivo comprimido en su misma carpeta.
    Retorna (num_archivos_extraidos, error_o_None).
    """
    archive_path = Path(archive_path)
    ext = archive_path.suffix.lower()
    extract_dir = _extract_dir_for_archive(archive_path)

    try:
        if ext == ".zip" or ext == ".kmz":
            os.makedirs(_long_path(extract_dir), exist_ok=True)
            with zipfile.ZipFile(_long_path(archive_path), "r") as zf:
                zf.extractall(_long_path(extract_dir))
                return (len(zf.namelist()), None)

        elif ext == ".rar":
            # Usar UnRAR directamente para mejor soporte de paths largos
            return _extract_rar_with_unrar(archive_path, extract_dir)

        elif ext == ".7z":
            if not PY7ZR_DISPONIBLE:
                return (0, "py7zr no instalado (pip install py7zr)")
            os.makedirs(_long_path(extract_dir), exist_ok=True)
            with py7zr.SevenZipFile(_long_path(archive_path), "r") as sz:
                sz.extractall(_long_path(extract_dir))
                return (len(sz.getnames()), None)

        else:
            return (0, f"Formato no soportado: {ext}")

    except Exception as e:
        return (0, str(e))


def _archive_entry_key(entry):
    """Clave estable para deduplicar archivos comprimidos ya procesados."""
    return _path_key(entry["path"])


def _extract_dir_for_archive(archive_path):
    """Retorna carpeta destino de extraccion para un archivo comprimido."""
    archive_path = Path(archive_path)
    return archive_path.parent / compact_component(
        archive_path.stem, maxlen=MAX_EXTRACT_DIR_LEN
    )


def _register_tree_levels(root_dir, level, levels_map):
    """Registra nivel de descompresion para todos los archivos bajo un directorio."""
    root_dir = Path(root_dir)
    if not _path_exists(root_dir):
        return
    for root, _, files in _walk_paths(root_dir):
        for f in files:
            p = Path(root) / f
            key = _path_key(p)
            prev = levels_map.get(key)
            if prev is None or level < prev:
                levels_map[key] = level


def _task_entry_keys(task):
    """Obtiene claves de los archivos comprimidos implicados en una tarea."""
    if task["kind"] == "single":
        return [_archive_entry_key(task["entry"])]
    keys = []
    for item in task.get("volumes", []):
        keys.append(_archive_entry_key(item))
    if not keys:
        keys.append(_archive_entry_key(task["start_item"]))
    return keys


def extract_all_archives(output_dir, documents=None, recursive=True, max_passes=MAX_EXTRACT_PASSES):
    """
    Extrae archivos comprimidos en output_dir.
    - Primera pasada: usa `documents` (si se entrega) para acotar a esta ejecucion.
    - Pasadas siguientes: descubre comprimidos nuevos (anidados) en todo output_dir.
    Retorna metricas de extraccion y mapa de niveles de descompresion.
    """
    output_dir = Path(output_dir).resolve()
    extracted_ok = 0
    extracted_fail = 0
    total_files_extracted = 0
    processed_keys = set()
    extraction_levels = {}
    pass_num = 0
    seed_docs = [
        d for d in (documents or [])
        if d.get("status") == "done" and d.get("download_path")
    ]

    # Base de niveles confiable: documentos descargados (status=done).
    # Esto evita etiquetar como "nivel 0" archivos heredados de corridas previas.
    for doc in seed_docs:
        rel_path = doc["download_path"]
        filepath = output_dir / rel_path
        if not _path_exists(filepath):
            continue
        extraction_levels.setdefault(_path_key(filepath), 0)

    while True:
        source_docs = seed_docs if (pass_num == 0 and seed_docs) else None
        entries = _discover_archive_entries(output_dir, documents=source_docs)
        pending = [e for e in entries if _archive_entry_key(e) not in processed_keys]

        if not pending:
            if pass_num == 0:
                print("  No se encontraron archivos comprimidos.")
            break

        pass_num += 1
        if pass_num > max_passes:
            print(f"  Se alcanzo maximo de pasadas de extraccion ({max_passes}).")
            break

        print(f"  Pasada {pass_num}: {len(pending)} comprimidos pendientes.")
        tasks = _build_extraction_plan(pending, output_dir)
        total = len(tasks)
        if len(pending) != len(tasks):
            print(f"  Se consolidaron {len(pending)} archivos en {len(tasks)} tareas "
                  f"(agrupando volumenes RAR).")

        for i, task in enumerate(tasks):
            for key in _task_entry_keys(task):
                processed_keys.add(key)
                extraction_levels.setdefault(key, 0)

            if task["kind"] == "single":
                archive_level = extraction_levels.get(_archive_entry_key(task["entry"]), 0)
            else:
                archive_level = extraction_levels.get(_archive_entry_key(task["start_item"]), 0)

            if task["kind"] == "single":
                archive = task["entry"]["path"]
                try:
                    rel = archive.relative_to(output_dir)
                    name_display = str(rel)
                except ValueError:
                    name_display = str(archive)
            else:
                start_item = task["start_item"]
                try:
                    rel = start_item["path"].relative_to(output_dir)
                    name_display = f"{rel} [vols:{len(task['volumes'])}]"
                except ValueError:
                    name_display = f"{start_item['path']} [vols:{len(task['volumes'])}]"

            if len(name_display) > 60:
                name_display = "..." + name_display[-57:]

            print(f"  [{i+1}/{total}] {console_safe(name_display)} ... ", end="", flush=True)

            if task["kind"] == "single":
                num_files, error = extract_archive(task["entry"]["path"])
                target_extract_dir = _extract_dir_for_archive(task["entry"]["path"])
            else:
                if not task["has_part1"]:
                    num_files, error = (0, "Falta volumen inicial .part1.rar")
                else:
                    start_name = task["start_item"]["source_name"]
                    num_files, error = _extract_rar_group_with_unrar(
                        task["volumes"], start_name, task["extract_dir"]
                    )
                target_extract_dir = task["extract_dir"]

            if error:
                extracted_fail += 1
                print(f"ERROR ({error})")
            else:
                extracted_ok += 1
                total_files_extracted += num_files
                _register_tree_levels(
                    target_extract_dir,
                    level=archive_level + 1,
                    levels_map=extraction_levels,
                )
                print(f"OK ({num_files} archivos)")

        if not recursive:
            break

    return extracted_ok, extracted_fail, total_files_extracted, extraction_levels


def _safe_stat_size(filepath):
    """Obtiene tamano de archivo manejando paths largos en Windows."""
    try:
        return filepath.stat().st_size
    except (OSError, FileNotFoundError):
        try:
            # Intentar con prefijo de path largo
            long_path = _long_path(filepath)
            return os.stat(long_path).st_size
        except (OSError, FileNotFoundError):
            return 0


def _path_exists(filepath):
    """Verifica existencia de path, incluyendo rutas largas en Windows."""
    p = Path(filepath)
    if p.exists():
        return True
    try:
        return os.path.exists(_long_path(p))
    except OSError:
        return False


def repair_downloaded_file_extensions(output_dir, documents):
    """
    Corrige extensiones de archivos ya descargados cuando quedaron con pseudo-extensiones
    por el nombre del enlace (ej: ".1 antecedentes legales") en vez de su extension real.
    """
    if not documents:
        return 0

    output_dir = Path(output_dir).resolve()
    fixed = 0

    for doc in documents:
        if doc.get("status") != "done":
            continue
        rel_path = doc.get("download_path")
        if not rel_path:
            continue

        expected_ext = normalize_known_extension(doc.get("file_type", ""))
        if not expected_ext:
            continue

        current_path = output_dir / rel_path
        if not _path_exists(current_path):
            continue
        if current_path.suffix.lower() == expected_ext:
            continue

        new_path = current_path.with_suffix(expected_ext)
        if _path_exists(new_path):
            # Evitar colision manteniendo nombre estable
            counter = 2
            while True:
                candidate = new_path.with_name(f"{new_path.stem}_{counter}{expected_ext}")
                if not _path_exists(candidate):
                    new_path = candidate
                    break
                counter += 1

        try:
            os.replace(_long_path(current_path), _long_path(new_path))
            doc["download_path"] = str(new_path.relative_to(output_dir))
            fixed += 1
        except OSError:
            pass

    return fixed


def collect_file_metrics(output_dir):
    """
    Recorre la carpeta y genera metricas completas de todos los archivos.
    Retorna dict con metricas.
    """
    type_counter = Counter()
    type_size = Counter()
    total_files = 0
    total_size = 0
    files_by_section = Counter()

    output_dir = Path(output_dir).resolve()
    for root, dirs, files in _walk_paths(output_dir):
        for f in files:
            if f == "manifest.json":
                continue
            filepath = Path(root) / f
            size = _safe_stat_size(filepath)
            ext = filepath.suffix.lower() or "(sin ext)"

            # Seccion = primer subdirectorio relativo a output_dir
            try:
                rel = filepath.relative_to(output_dir)
                section = rel.parts[0] if len(rel.parts) > 1 else "(raiz)"
            except ValueError:
                section = "(otro)"

            type_counter[ext] += 1
            type_size[ext] += size
            files_by_section[section] += 1
            total_files += 1
            total_size += size

    return {
        "total_files": total_files,
        "total_size": total_size,
        "by_type": type_counter,
        "by_type_size": type_size,
        "by_section": files_by_section,
    }


def print_extraction_metrics(output_dir):
    """Imprime metricas detalladas post-extraccion."""
    metrics = collect_file_metrics(output_dir)

    print()
    print("=" * 64)
    print("METRICAS POST-EXTRACCION")
    print("=" * 64)
    print(f"  Total archivos en carpeta:  {metrics['total_files']}")
    print(f"  Tamano total:               {format_size(metrics['total_size'])}")

    print()
    print("  Por tipo de archivo:")
    print(f"    {'Extension':12s} {'Cantidad':>10s} {'Tamano':>14s}")
    print(f"    {'-'*12} {'-'*10} {'-'*14}")
    for ext, count in metrics["by_type"].most_common():
        size = format_size(metrics["by_type_size"][ext])
        print(f"    {ext:12s} {count:>10d} {size:>14s}")

    print()
    print("  Por seccion:")
    print(f"    {'Seccion':50s} {'Archivos':>10s}")
    print(f"    {'-'*50} {'-'*10}")
    for section, count in metrics["by_section"].most_common():
        sec_display = section if len(section) <= 50 else section[:47] + "..."
        print(f"    {sec_display:50s} {count:>10d}")


def collect_documents_for_report(output_dir, extensions=REPORT_EXTENSIONS):
    """
    Recolecta archivos para reporte final.
    Por defecto incluye .pdf y .doc.
    Si `extensions` es None, incluye todos los tipos.
    """
    extensions_filter = None
    if extensions is not None:
        extensions_filter = {str(ext).lower() for ext in extensions}

    output_dir = Path(output_dir).resolve()
    docs = []

    for root, _, files in _walk_paths(output_dir):
        for f in files:
            if f in GENERATED_LOCAL_FILENAMES:
                continue
            if f.startswith("~$"):
                continue
            filepath = Path(root) / f
            ext = filepath.suffix.lower()
            if extensions_filter is not None and ext not in extensions_filter:
                continue

            try:
                rel_path = filepath.relative_to(output_dir)
            except ValueError:
                rel_path = filepath

            docs.append({
                "nombre_archivo": filepath.name,
                "extension": ext,
                "ruta_relativa": str(rel_path),
                "ruta_absoluta": _safe_resolve_str(filepath),
                "tamano_bytes": _safe_stat_size(filepath),
            })

    docs.sort(key=lambda x: (x["extension"], x["ruta_relativa"].lower()))
    return docs


def collect_downloaded_documents_for_upload(output_dir, documents=None, include_archives=False):
    """
    Recolecta archivos para CP8 luego de la extraccion.
    Por defecto excluye archivos comprimidos para subir contenido descomprimido.
    """
    output_dir = Path(output_dir).resolve()
    items = []

    for root, _, files in _walk_paths(output_dir):
        for f in files:
            if f in GENERATED_LOCAL_FILENAMES:
                continue
            if f.startswith("~$"):
                continue
            filepath = Path(root) / f
            ext = filepath.suffix.lower()
            if not include_archives and ext in ARCHIVE_EXTENSIONS:
                continue

            try:
                rel_path = filepath.relative_to(output_dir)
            except ValueError:
                rel_path = filepath

            items.append({
                "nombre_archivo": filepath.name,
                "extension": ext,
                "ruta_relativa": str(rel_path),
                "ruta_absoluta": _safe_resolve_str(filepath),
                "tamano_bytes": _safe_stat_size(filepath),
            })

    items.sort(key=lambda x: (x["extension"], x["ruta_relativa"].lower()))
    return items


def _normalize_match_text(text):
    """Normaliza texto para comparacion robusta (sin tildes, minusculas)."""
    s = clean_text(text or "")
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    return s.lower()


def build_final_pdf_report_from_trace(trace_rows, exclude_keywords=FINAL_PDF_EXCLUDE_KEYWORDS):
    """
    Construye listado final de PDF a partir de trazabilidad.
    Reglas:
    - incluir SIEMPRE PDF con nivel_descarga_descompresion == 0
    - para PDF con nivel > 0, excluir si nombre_archivo contiene palabras bloqueadas
      (sin distinguir mayusculas/minusculas ni tildes).
    """
    raw_keywords = [k for k in (exclude_keywords or []) if k]
    keywords_norm = [(_normalize_match_text(k), k) for k in raw_keywords]
    docs = []
    stats = {
        "trace_total_files": 0,
        "descartados_extension_no_pdf": 0,
        "descartados_extension_detalle": Counter(),
        "pdf_totales_trazabilidad": 0,
        "pdf_nivel_0_incluidos": 0,
        "pdf_nivel_mayor_0_candidatos": 0,
        "pdf_nivel_mayor_0_incluidos": 0,
        "pdf_excluidos_por_palabra": 0,
        "pdf_excluidos_por_palabra_detalle": {k: 0 for k in raw_keywords},
        "pdf_finales": 0,
    }

    for row in trace_rows or []:
        stats["trace_total_files"] += 1
        ext = str(row.get("extension") or "").lower()
        if ext != ".pdf":
            stats["descartados_extension_no_pdf"] += 1
            stats["descartados_extension_detalle"][ext or "(sin ext)"] += 1
            continue
        stats["pdf_totales_trazabilidad"] += 1

        level_raw = row.get("nivel_descarga_descompresion")
        try:
            level = int(level_raw)
        except (TypeError, ValueError):
            level = 0

        file_name = str(
            row.get("nombre_archivo")
            or Path(str(row.get("ruta_relativa") or row.get("ruta") or "")).name
        )
        file_name_norm = _normalize_match_text(file_name)
        matched_keyword_raw = None
        for kw_norm, kw_raw in keywords_norm:
            if kw_norm and kw_norm in file_name_norm:
                matched_keyword_raw = kw_raw
                break

        if level == 0:
            stats["pdf_nivel_0_incluidos"] += 1
        else:
            stats["pdf_nivel_mayor_0_candidatos"] += 1

        if level > 0 and matched_keyword_raw is not None:
            stats["pdf_excluidos_por_palabra"] += 1
            stats["pdf_excluidos_por_palabra_detalle"][matched_keyword_raw] += 1
            continue
        if level > 0:
            stats["pdf_nivel_mayor_0_incluidos"] += 1

        size_bytes = row.get("tamano_bytes")
        if size_bytes in (None, ""):
            size_bytes = row.get("peso_bytes") or 0

        docs.append({
            "tipo": "",
            "seleccionar": True,
            "nombre_archivo": file_name,
            "nombre_archivo_final": file_name,
            "extension": ".pdf",
            "formato": "pdf",
            "ruta_relativa": str(row.get("ruta_relativa") or row.get("ruta") or ""),
            "ruta_absoluta": str(row.get("ruta_absoluta") or ""),
            "ruta_local_final": str(row.get("ruta_absoluta") or ""),
            "tamano_bytes": size_bytes,
            "nivel_descarga_descompresion": level,
            "origen": str(row.get("origen") or ""),
            "categoria": str(row.get("categoria") or ""),
            "texto_link": str(row.get("texto_link") or ""),
            "url_origen": str(row.get("url_origen") or ""),
            "archivo_descargado_raiz": str(row.get("archivo_descargado_raiz") or ""),
            "ruta_archivo_descargado_raiz": str(row.get("ruta_archivo_descargado_raiz") or ""),
            "selected": True,
            "upload_status": "pending",
        })

    docs.sort(key=lambda x: (x["nivel_descarga_descompresion"], x["ruta_relativa"].lower()))
    stats["pdf_finales"] = len(docs)
    return docs, stats


def build_semantic_final_filename(item, parent_dir):
    """Construye nombre final semantico para el PDF filtrado."""
    current_name = str(item.get("nombre_archivo") or "")
    current_path = Path(item.get("ruta_relativa") or current_name)
    ext = current_path.suffix.lower() or ".pdf"
    stem_source = current_path.stem or Path(current_name).stem or "archivo"
    parts = [
        clean_text(item.get("tipo") or ""),
        clean_text(item.get("categoria") or ""),
        clean_text(item.get("texto_link") or ""),
        clean_text(stem_source),
    ]
    semantic_stem = "_".join([p for p in parts if p]) or stem_source
    stem_budget = compute_stem_budget(parent_dir, suffix_len=len(ext), hard_cap=MAX_FILENAME_LEN)
    return compact_component(semantic_stem, maxlen=stem_budget) + ext


def build_notebook_upload_filename(item):
    """Construye nombre descriptivo para mostrar/subir al notebook sin depender del nombre local."""
    original_name = str(item.get("nombre_archivo") or "")
    final_name = str(item.get("nombre_archivo_final") or "")
    current_path = Path(item.get("ruta_relativa") or final_name or original_name)
    ext = (
        Path(final_name).suffix.lower()
        or Path(original_name).suffix.lower()
        or current_path.suffix.lower()
        or ".pdf"
    )
    stem_source = (
        Path(original_name).stem
        or Path(final_name).stem
        or current_path.stem
        or "archivo"
    )
    texto_link = clean_text(item.get("texto_link") or "")

    # Si el nombre original ya fue compactado con hash al descargarlo,
    # preferimos el texto del link para reconstruir un nombre legible.
    stem_compact = compact_component(stem_source, maxlen=500)
    texto_link_compact = compact_component(texto_link, maxlen=500) if texto_link else ""
    if re.search(r"_[0-9a-f]{8}$", stem_compact):
        stem_without_hash = re.sub(r"_[0-9a-f]{8}$", "", stem_compact)
        if texto_link_compact and (
            texto_link_compact.startswith(stem_without_hash)
            or stem_without_hash.startswith(texto_link_compact[: max(1, len(stem_without_hash))])
        ):
            stem_source = texto_link

    parts = [
        clean_text(item.get("tipo") or ""),
        clean_text(item.get("categoria") or ""),
        texto_link,
        clean_text(stem_source),
    ]
    semantic_stem = "_".join([p for p in parts if p]) or stem_source
    normalized = str(semantic_stem or "archivo")
    for char in '<>:"/\\|?*':
        normalized = normalized.replace(char, "_")
    normalized = unicodedata.normalize("NFKD", normalized)
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    normalized = normalized.replace(".", "_")
    normalized = re.sub(r"[ _]+", "_", normalized).strip("_")
    if not normalized:
        normalized = "archivo"
    return normalized + ext


def rename_final_pdf_documents(output_dir, docs):
    """
    Renombra fisicamente los PDF finales usando contexto semantico.
    Actualiza in-place nombre_archivo_final, ruta_relativa y ruta_local_final.
    """
    output_dir = Path(output_dir).resolve()
    renamed = 0

    for item in docs or []:
        current_rel = str(item.get("ruta_relativa") or "")
        if not current_rel:
            item["nombre_archivo_final"] = item.get("nombre_archivo") or ""
            item["nombre_archivo_notebook"] = build_notebook_upload_filename(item)
            continue

        current_path = output_dir / current_rel
        if not _path_exists(current_path):
            item["nombre_archivo_final"] = item.get("nombre_archivo") or Path(current_rel).name
            item["nombre_archivo_notebook"] = build_notebook_upload_filename(item)
            continue

        parent_dir = current_path.parent
        target_name = build_semantic_final_filename(item, parent_dir)
        target_path = parent_dir / target_name

        if _path_key(target_path) != _path_key(current_path):
            counter = 2
            while _path_exists(target_path):
                extra = f"_{counter}{target_path.suffix.lower()}"
                stem_budget = compute_stem_budget(parent_dir, suffix_len=len(extra), hard_cap=MAX_FILENAME_LEN)
                base_stem = build_semantic_final_filename(item, parent_dir)
                base_stem = Path(base_stem).stem
                dedup_name = f"{compact_component(base_stem, maxlen=stem_budget)}_{counter}{target_path.suffix.lower()}"
                target_path = parent_dir / dedup_name
                counter += 1

            try:
                os.replace(_long_path(current_path), _long_path(target_path))
                renamed += 1
                current_path = target_path
            except OSError:
                current_path = output_dir / current_rel

        item["nombre_archivo_final"] = current_path.name
        item["ruta_relativa"] = str(current_path.relative_to(output_dir))
        item["ruta_absoluta"] = _safe_resolve_str(current_path)
        item["ruta_local_final"] = item["ruta_absoluta"]
        item["formato"] = current_path.suffix.lower().replace(".", "") or "pdf"
        item["nombre_archivo_notebook"] = build_notebook_upload_filename(item)

    return renamed


def print_final_pdf_report(docs, stats):
    """Imprime listado final PDF y resumen de filtros aplicados."""
    print()
    print("=" * 64)
    print("LISTADO FINAL PDF (CP6B)")
    print("=" * 64)
    print(f"  PDF en trazabilidad:            {stats.get('pdf_totales_trazabilidad', 0)}")
    print(f"  PDF nivel 0 incluidos:          {stats.get('pdf_nivel_0_incluidos', 0)}")
    print(f"  PDF nivel > 0 candidatos:       {stats.get('pdf_nivel_mayor_0_candidatos', 0)}")
    print(f"  PDF nivel > 0 incluidos:        {stats.get('pdf_nivel_mayor_0_incluidos', 0)}")
    print(f"  PDF excluidos por palabra:      {stats.get('pdf_excluidos_por_palabra', 0)}")
    print(f"  TOTAL LISTADO FINAL PDF:        {len(docs)}")
    if not docs:
        print("  No se encontraron PDF para listado final.")
        return

    for i, item in enumerate(docs, 1):
        print(
            f"  [{i}] L{item['nivel_descarga_descompresion']} | "
            f"{console_safe(item['ruta_relativa'])}"
        )


def export_final_pdf_report_excel(docs, output_dir):
    """
    Exporta listado final PDF a Excel.
    Retorna (excel_path, error_o_None).
    """
    output_dir = Path(output_dir)
    excel_path = output_dir / "listado_final_pdf.xlsx"

    try:
        from openpyxl import Workbook
    except ImportError:
        return (None, "No se pudo generar Excel: openpyxl no instalado (pip install openpyxl)")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Listado Final PDF"

        headers = [
            "tipo",
            "seleccionar",
            "categoria",
            "texto_link",
            "url_origen",
            "nombre_archivo",
            "nombre_archivo_final",
            "nombre_archivo_notebook",
            "extension",
            "formato",
            "ruta_relativa",
            "ruta_local_final",
            "ruta_absoluta",
            "tamano_bytes",
            "tamano_legible",
            "nivel_descarga_descompresion",
            "origen",
            "archivo_descargado_raiz",
            "ruta_archivo_descargado_raiz",
        ]
        ws.append(headers)

        for item in docs:
            ws.append([
                item.get("tipo"),
                item.get("seleccionar"),
                item.get("categoria"),
                item.get("texto_link"),
                item.get("url_origen"),
                item["nombre_archivo"],
                item.get("nombre_archivo_final"),
                item.get("nombre_archivo_notebook"),
                item["extension"],
                item.get("formato"),
                item["ruta_relativa"],
                item.get("ruta_local_final"),
                item["ruta_absoluta"],
                item["tamano_bytes"],
                format_size(item["tamano_bytes"]),
                item.get("nivel_descarga_descompresion"),
                item.get("origen"),
                item.get("archivo_descargado_raiz"),
                item.get("ruta_archivo_descargado_raiz"),
            ])

        column_widths = {
            "A": 20,
            "B": 12,
            "C": 28,
            "D": 40,
            "E": 70,
            "F": 40,
            "G": 40,
            "H": 48,
            "I": 12,
            "J": 12,
            "K": 90,
            "L": 120,
            "M": 120,
            "N": 15,
            "O": 15,
            "P": 24,
            "Q": 16,
            "R": 40,
            "S": 90,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(excel_path)
        return (excel_path, None)
    except Exception as e:
        return (None, f"No se pudo generar Excel: {e}")


def print_post_filter_summary(initial_type_counter, post_metrics, filter_stats, exclude_keywords=None):
    """
    Imprime resumen final de consistencia luego de generar el listado filtrado.
    """
    initial_type_counter = initial_type_counter or Counter()
    post_metrics = post_metrics or {}
    post_by_type = post_metrics.get("by_type") or Counter()

    print()
    print("=" * 64)
    print("RESUMEN FINAL DE FILTRO")
    print("=" * 64)

    print("  1) Archivos al comienzo (CP2) por tipo:")
    if not initial_type_counter:
        print("    (sin datos)")
    else:
        for ext, count in initial_type_counter.most_common():
            ext_display = ext if ext else "(sin ext)"
            print(f"    {ext_display:12s} {count:>8d}")

    print()
    print("  2) Archivos despues de descomprimir (carpeta final) por tipo:")
    if not post_by_type:
        print("    (sin datos)")
    else:
        for ext, count in post_by_type.most_common():
            ext_display = ext if ext else "(sin ext)"
            print(f"    {ext_display:12s} {count:>8d}")

    print()
    print("  3) Descartados por extension (no .pdf):")
    print(f"    Total descartados: {filter_stats.get('descartados_extension_no_pdf', 0)}")

    print()
    print("  4) Descartados por filtro de palabras (por palabra):")
    keyword_detail = filter_stats.get("pdf_excluidos_por_palabra_detalle") or {}
    for keyword in (exclude_keywords if exclude_keywords is not None else FINAL_PDF_EXCLUDE_KEYWORDS):
        print(f"    {keyword:12s} {int(keyword_detail.get(keyword, 0)):>8d}")

    print()
    print("  5) Total de archivos PDF que se cargaran:")
    print(f"    TOTAL PDF FINAL: {filter_stats.get('pdf_finales', 0)}")


def print_documents_report(docs):
    """Imprime listado de archivos PDF y DOC con su ruta."""
    print()
    print("=" * 64)
    print("LISTADO FINAL PDF/DOC")
    print("=" * 64)
    print(f"  Total archivos: {len(docs)}")
    if not docs:
        print("  No se encontraron archivos .pdf o .doc")
        return

    for i, item in enumerate(docs, 1):
        print(f"  [{i}] {item['extension']} | {console_safe(item['ruta_relativa'])}")


def export_documents_report_excel(docs, output_dir):
    """
    Exporta listado de documentos PDF/DOC a Excel.
    Retorna (excel_path, error_o_None).
    """
    output_dir = Path(output_dir)
    excel_path = output_dir / "listado_pdf_doc.xlsx"

    try:
        from openpyxl import Workbook
    except ImportError:
        return (None, "No se pudo generar Excel: openpyxl no instalado (pip install openpyxl)")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Listado PDF DOC"

        headers = [
            "nombre_archivo",
            "extension",
            "ruta_relativa",
            "ruta_absoluta",
            "tamano_bytes",
            "tamano_legible",
        ]
        ws.append(headers)

        for item in docs:
            ws.append([
                item["nombre_archivo"],
                item["extension"],
                item["ruta_relativa"],
                item["ruta_absoluta"],
                item["tamano_bytes"],
                format_size(item["tamano_bytes"]),
            ])

        # Ajuste simple de ancho de columnas
        column_widths = {
            "A": 45,
            "B": 10,
            "C": 80,
            "D": 110,
            "E": 15,
            "F": 15,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(excel_path)
        return (excel_path, None)
    except Exception as e:
        return (None, f"No se pudo generar Excel: {e}")


# ============================================================================
# MAIN
# ============================================================================

def parse_args():
    """Parsea argumentos o pide URL interactivamente."""
    parser = argparse.ArgumentParser(
        description="Descarga documentos de una pagina SEIA",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Ejemplo:\n  python download_documento_seia.py "
               '"https://seia.sea.gob.cl/documentos/documento.php?idDocumento=2141618022"'
    )
    parser.add_argument("url", nargs="?", help="URL de la pagina de documento SEIA")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT,
                        help=f"Carpeta de salida (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--skip-size-estimation", action="store_true",
                        help="Omitir estimacion de tamanos (mas rapido)")
    parser.add_argument("--no-extract", action="store_true",
                        help="No extraer archivos comprimidos despues de descargar")
    parser.add_argument("--keep-existing", action="store_true",
                        help="No limpiar carpeta de salida al iniciar (default: limpiar)")

    args = parser.parse_args()

    if not args.url:
        print()
        args.url = input("Ingresa la URL del documento SEIA: ").strip()
        if not args.url:
            print("Error: No se proporciono una URL.")
            sys.exit(1)

    return args


def validate_url(url):
    """Valida que la URL sea de documento SEIA."""
    parsed = urlparse(url)
    if not parsed.scheme:
        url = "https://" + url
        parsed = urlparse(url)
    if "seia.sea.gob.cl" not in parsed.netloc:
        print(f"Advertencia: La URL no parece ser del SEIA ({parsed.netloc}).")
        print("Se intentara de todas formas.")
    return url


def extract_id_documento(url):
    """Extrae el idDocumento de la URL."""
    parsed = urlparse(url)
    if parsed.query:
        for key, value in parse_qsl(parsed.query, keep_blank_values=True):
            key_l = key.lower()
            if key_l in {"iddocumento", "docid"} and value:
                return value.strip("/")
    return "desconocido"


def prepare_output_directory(output_dir, base_output_dir, clean_start=True):
    """
    Prepara carpeta de salida.
    Si clean_start=True, elimina la carpeta objetivo para evitar duplicados.
    """
    output_dir = Path(output_dir).resolve()
    base_output_dir = Path(base_output_dir).resolve()

    if clean_start and output_dir.exists():
        try:
            output_dir.relative_to(base_output_dir)
        except ValueError:
            raise RuntimeError(
                f"Seguridad: no se limpiara '{output_dir}' porque no cuelga de '{base_output_dir}'."
            )
        if output_dir == base_output_dir:
            raise RuntimeError(
                f"Seguridad: no se limpiara la carpeta base completa '{base_output_dir}'."
            )
        shutil.rmtree(output_dir, ignore_errors=True)

    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


class NotebookAPIError(RuntimeError):
    """Error al interactuar con la API de notebook."""


def build_notebook_title():
    """Construye el titulo de notebook con fecha-hora local."""
    return f"Notebook {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"


def build_adenda_notebook_title(id_adenda, id_documento):
    """Construye nombre consistente para notebook de una adenda."""
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe_doc_id = compact_component(str(id_documento or "desconocido"), maxlen=60)
    return f"ADENDA-{id_adenda}_DOC-{safe_doc_id}_{stamp}"


def build_tipo_notebook_title(tipo, id_documento):
    """Construye nombre consistente para notebook basado en tipo."""
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    tipo_slug = compact_component(tipo or "sin_tipo", maxlen=40)
    safe_doc_id = compact_component(str(id_documento or "desconocido"), maxlen=60)
    return f"{tipo_slug}_DOC-{safe_doc_id}_{stamp}"


def _source_payload(source):
    """Convierte objeto Source de notebooklm-py a dict JSON."""
    return {
        "id": source.id,
        "title": source.title,
        "kind": str(source.kind),
        "status": int(source.status),
        "status_text": source_status_to_str(source.status),
    }


def _notebooklm_error_message(exc):
    """Mapea excepciones de notebooklm-py a mensajes legibles."""
    if isinstance(exc, AuthError):
        return f"Error de autenticacion NotebookLM: {exc}"
    if isinstance(exc, SourceTimeoutError):
        return f"Timeout procesando source: {exc}"
    if isinstance(exc, SourceProcessingError):
        return f"Error procesando source: {exc}"
    if isinstance(exc, RPCError):
        return f"Error RPC NotebookLM: {exc}"
    return f"Error NotebookLM: {exc}"


def _normalize_notebook_auth_payload(
    notebook_auth: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    """Normaliza el payload compacto de auth recibido por header."""
    if not notebook_auth:
        return None

    if not isinstance(notebook_auth, dict):
        raise NotebookAPIError("El payload de auth de NotebookLM no tiene un formato valido.")

    raw_cookies = notebook_auth.get("cookies")
    if not isinstance(raw_cookies, dict):
        raise NotebookAPIError("El payload de auth de NotebookLM debe incluir 'cookies'.")

    cookies: Dict[str, str] = {}
    for raw_name, raw_value in raw_cookies.items():
        if raw_value is None:
            continue
        name = str(raw_name).strip()
        value = str(raw_value)
        if name and value:
            cookies[name] = value

    if "SID" not in cookies:
        raise NotebookAPIError("El payload de auth de NotebookLM no incluye la cookie SID.")

    return {
        "version": 1,
        "cookies": cookies,
        "cookie_names": sorted(cookies.keys()),
        "cookie_domains": [
            str(domain).strip()
            for domain in notebook_auth.get("cookie_domains", [])
            if str(domain).strip()
        ],
    }


def prepare_notebook_client_seed(
    notebook_auth: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    """Prepara cookies y tokens una sola vez para reusar el auth per-request."""
    normalized_auth = _normalize_notebook_auth_payload(notebook_auth)
    if not normalized_auth:
        return None

    async def _prepare():
        csrf_token, session_id = await fetch_tokens(normalized_auth["cookies"])
        return {
            "cookies": dict(normalized_auth["cookies"]),
            "csrf_token": csrf_token,
            "session_id": session_id,
        }

    return asyncio.run(_prepare())


def _auth_tokens_from_seed(auth_seed: Dict[str, Any]) -> AuthTokens:
    """Construye AuthTokens frescos desde un seed ya validado."""
    return AuthTokens(
        cookies=dict(auth_seed["cookies"]),
        csrf_token=str(auth_seed["csrf_token"]),
        session_id=str(auth_seed["session_id"]),
    )


async def _create_notebook_client_async(
    *,
    notebook_auth: Optional[Dict[str, Any]] = None,
    auth_seed: Optional[Dict[str, Any]] = None,
    timeout: float = NOTEBOOK_CLIENT_TIMEOUT_SEC,
) -> NotebookLMClient:
    """Crea cliente NotebookLM usando auth per-request o fallback a storage."""
    if auth_seed:
        return NotebookLMClient(_auth_tokens_from_seed(auth_seed), timeout=timeout)

    normalized_auth = _normalize_notebook_auth_payload(notebook_auth)
    if normalized_auth:
        csrf_token, session_id = await fetch_tokens(normalized_auth["cookies"])
        return NotebookLMClient(
            AuthTokens(
                cookies=dict(normalized_auth["cookies"]),
                csrf_token=csrf_token,
                session_id=session_id,
            ),
            timeout=timeout,
        )

    return await NotebookLMClient.from_storage(timeout=timeout)


def notify_notebook_api(
    notebook_title=None,
    api_base_url=None,
    raise_on_error=False,
    notebook_auth=None,
    auth_seed=None,
):
    """Crea un notebook via notebooklm-py directamente."""
    title = notebook_title or build_notebook_title()
    print()
    print("[CP7] Creando notebook en NotebookLM...")
    print(f"  Titulo: {title}")

    notebook_id = None
    error = None
    try:
        async def _create():
            async with await _create_notebook_client_async(
                notebook_auth=notebook_auth,
                auth_seed=auth_seed,
                timeout=NOTEBOOK_CLIENT_TIMEOUT_SEC,
            ) as client:
                nb = await client.notebooks.create(title)
            return nb.id, nb.title

        notebook_id, created_title = asyncio.run(_create())
        title = created_title or title
        print(f"  Notebook creado: id={notebook_id}")
    except Exception as e:
        error = _notebooklm_error_message(e)
        print(f"  {error}")

    if not notebook_id and not error:
        error = "No se obtuvo notebook_id"

    if error and raise_on_error:
        raise NotebookAPIError(error)
    return notebook_id, title, error


def list_notebook_sources(notebook_id, api_base_url=None, notebook_auth=None, auth_seed=None):
    """Lista fuentes de un notebook via notebooklm-py directamente."""
    async def _list():
        async with await _create_notebook_client_async(
            notebook_auth=notebook_auth,
            auth_seed=auth_seed,
            timeout=NOTEBOOK_CLIENT_TIMEOUT_SEC,
        ) as client:
            sources = await client.sources.list(notebook_id)
        return [_source_payload(s) for s in sources]

    return asyncio.run(_list())


def select_first_documents_for_upload(docs_report, limit=None):
    """
    Selecciona primeros documentos para carga:
    prioriza PDF, y completa con DOC si faltan.
    """
    existing_docs = list(docs_report)

    if limit is None:
        return existing_docs

    pdf_docs = [
        d for d in existing_docs
        if d.get("extension") == ".pdf"
    ]
    doc_docs = [
        d for d in existing_docs
        if d.get("extension") == ".doc"
    ]
    other_docs = [
        d for d in existing_docs
        if d.get("extension") not in {".pdf", ".doc"}
    ]

    selected = pdf_docs[:limit]
    if len(selected) < limit:
        selected.extend(doc_docs[:limit - len(selected)])
    if len(selected) < limit:
        selected.extend(other_docs[:limit - len(selected)])
    return selected


def _upload_single_document(notebook_id, doc, order, notebook_auth=None, auth_seed=None):
    """Sube un solo documento al notebook via notebooklm-py directamente."""
    p = Path(doc["ruta_absoluta"])
    upload_name = str(
        doc.get("nombre_archivo_notebook") or build_notebook_upload_filename(doc)
    )
    base_item = {
        "_order": order,
        "document_id": doc.get("document_id"),
        "ruta_relativa": doc.get("ruta_relativa"),
        "nombre_archivo": doc.get("nombre_archivo") or p.name,
        "nombre_archivo_notebook": upload_name,
        "status_code": None,
        "uploaded": False,
        "error": None,
        "warning": None,
        "elapsed_seconds": 0.0,
    }

    if not _path_exists(p):
        return {**base_item, "error": f"No existe archivo: {p}", "attempts": 0}

    current_seed = auth_seed

    async def _upload():
        async with await _create_notebook_client_async(
            notebook_auth=notebook_auth,
            auth_seed=current_seed,
            timeout=NOTEBOOK_CLIENT_TIMEOUT_SEC,
        ) as client:
            source = await client.sources.add_file(
                notebook_id,
                _long_path(p),
                wait=True,
                wait_timeout=NOTEBOOK_UPLOAD_WAIT_TIMEOUT_SEC,
            )
            warning = None
            if upload_name and source.title != upload_name:
                try:
                    source = await client.sources.rename(
                        notebook_id,
                        source.id,
                        upload_name,
                    )
                except Exception as rename_exc:
                    warning = (
                        f"Source subido pero no se pudo renombrar a "
                        f"'{upload_name}': {_notebooklm_error_message(rename_exc)}"
                    )
        return source, warning

    async def _refresh_tokens():
        cookies = (current_seed or {}).get("cookies") or {}
        if not cookies:
            return None
        new_csrf, new_sess = await fetch_tokens(dict(cookies))
        return {
            **(current_seed or {}),
            "csrf_token": new_csrf,
            "session_id": new_sess,
        }

    started = time.perf_counter()
    last_error = None
    attempts_done = 0
    for attempt in range(1, NOTEBOOK_UPLOAD_RETRY_ATTEMPTS + 1):
        attempts_done = attempt
        try:
            source, warning = asyncio.run(_upload())
            elapsed = time.perf_counter() - started
            return {
                **base_item,
                "status_code": 200,
                "uploaded": True,
                "error": None,
                "warning": warning,
                "response_body": {"ok": True, "item": _source_payload(source)},
                "elapsed_seconds": round(elapsed, 2),
                "attempts": attempts_done,
            }
        except (FileNotFoundError, ValidationError, AuthError) as e:
            last_error = e
            break
        except httpx.HTTPStatusError as e:
            last_error = e
            status_code = getattr(getattr(e, "response", None), "status_code", 0) or 0
            if status_code != 429 and status_code < 500:
                break
        except (
            SourceAddError,
            SourceTimeoutError,
            SourceProcessingError,
            httpx.TimeoutException,
            httpx.TransportError,
        ) as e:
            last_error = e
        except Exception as e:
            last_error = e

        if attempt < NOTEBOOK_UPLOAD_RETRY_ATTEMPTS:
            refresh_note = ""
            if isinstance(last_error, SourceAddError) or "SOURCE_ID" in str(last_error):
                try:
                    refreshed = asyncio.run(_refresh_tokens())
                    if refreshed:
                        current_seed = refreshed
                        refresh_note = " [tokens refrescados]"
                except Exception as refresh_exc:
                    refresh_note = f" [refresh tokens fallo: {type(refresh_exc).__name__}]"
            sleep_s = NOTEBOOK_UPLOAD_RETRY_BASE_SEC * (2 ** (attempt - 1))
            sleep_s += random.uniform(0, 0.5)
            print(
                f"      Reintento {attempt}/{NOTEBOOK_UPLOAD_RETRY_ATTEMPTS - 1} "
                f"de {console_safe(upload_name)} en {round(sleep_s, 2)}s "
                f"(motivo: {type(last_error).__name__}){refresh_note}"
            )
            time.sleep(sleep_s)

    elapsed = time.perf_counter() - started
    error_msg = _notebooklm_error_message(last_error) if last_error else "Error desconocido"
    if attempts_done > 1:
        error_msg = f"{error_msg} [tras {attempts_done} intentos]"
    return {
        **base_item,
        "error": error_msg,
        "elapsed_seconds": round(elapsed, 2),
        "attempts": attempts_done,
    }


def upload_documents_batch_and_single(
    notebook_id,
    docs_report,
    limit=None,
    api_base_url=None,
    progress_callback=None,
    item_callback=None,
    notebook_auth=None,
    auth_seed=None,
):
    """
    Carga archivos al notebook en paralelo acotado usando notebooklm-py.
    Cada worker resuelve su propia corrida async para aprovechar la paralelizacion
    sin depender de un servidor HTTP intermedio.
    """
    selected = select_first_documents_for_upload(docs_report, limit=limit)
    if not selected:
        print()
        print("[CP8] No hay archivos para subir.")
        return {
            "selected": 0,
            "uploaded_ok": 0,
            "uploaded_failed": 0,
            "items": [],
        }

    print()
    print("[CP8] Subida de documentos al notebook")
    print(f"  Notebook ID: {notebook_id}")
    print(f"  Total objetivo: {len(selected)}")
    max_workers = max(1, min(NOTEBOOK_UPLOAD_MAX_WORKERS, len(selected)))
    if max_workers == 1:
        print("  Modo: secuencial (1 archivo a la vez)")
    else:
        print(f"  Modo: paralelo acotado ({max_workers} archivos a la vez)")
    print("  Archivos objetivo:")
    for i, doc in enumerate(selected, 1):
        print(f"    [{i}] {console_safe(doc['ruta_relativa'])}")

    print()
    if max_workers == 1:
        print("  Subida secuencial...")
    else:
        print(f"  Subida en paralelo con {max_workers} workers...")
    uploaded_ok = 0
    upload_items = []
    total_selected = len(selected)
    emit_progress(
        progress_callback,
        stage="uploading",
        current=0,
        total=total_selected,
        message=(
            f"Iniciando carga secuencial de {total_selected} documento(s)."
            if max_workers == 1
            else f"Iniciando carga paralela ({max_workers} workers) de {total_selected} documento(s)."
        ),
    )

    completed_count = 0
    future_map = {}
    resolved_auth_seed = auth_seed
    if resolved_auth_seed is None and notebook_auth is not None:
        resolved_auth_seed = prepare_notebook_client_seed(notebook_auth)

    def _notify_start(order, doc):
        p = Path(doc["ruta_absoluta"])
        upload_name = str(
            doc.get("nombre_archivo_notebook") or build_notebook_upload_filename(doc)
        )
        if item_callback:
            item_callback(
                "starting",
                {
                    "document_id": doc.get("document_id"),
                    "ruta_relativa": doc.get("ruta_relativa"),
                    "nombre_archivo": doc.get("nombre_archivo") or p.name,
                    "nombre_archivo_notebook": upload_name,
                    "uploaded": False,
                    "status_code": None,
                    "error": None,
                },
                order,
                total_selected,
            )

    def _submit_next(executor, pending_iterator):
        try:
            order, doc = next(pending_iterator)
        except StopIteration:
            return
        _notify_start(order, doc)
        if NOTEBOOK_UPLOAD_SUBMIT_JITTER_SEC > 0:
            time.sleep(random.uniform(0, NOTEBOOK_UPLOAD_SUBMIT_JITTER_SEC))
        future = executor.submit(
            _upload_single_document,
            notebook_id,
            doc,
            order,
            None,
            resolved_auth_seed,
        )
        future_map[future] = (order, doc)

    with cf.ThreadPoolExecutor(max_workers=max_workers) as executor:
        pending_iterator = iter(list(enumerate(selected, 1)))
        for _ in range(max_workers):
            _submit_next(executor, pending_iterator)

        while future_map:
            done, _ = cf.wait(
                tuple(future_map.keys()),
                return_when=cf.FIRST_COMPLETED,
            )
            for future in done:
                order, _doc = future_map.pop(future)
                item = future.result()
                completed_count += 1
                if item.get("uploaded"):
                    uploaded_ok += 1
                upload_items.append(item)

                upload_name = str(item.get("nombre_archivo_notebook") or "")
                if item.get("uploaded"):
                    print(
                        f"    [{order}] {console_safe(upload_name)} -> status "
                        f"{item.get('status_code')} ({item.get('elapsed_seconds')} s)"
                    )
                    if item.get("warning"):
                        print(f"      ADVERTENCIA: {console_safe(item.get('warning'))}")
                else:
                    print(
                        f"    [{order}] ERROR subiendo {console_safe(upload_name)}: "
                        f"{console_safe(item.get('error'))}"
                    )

                response_body = item.get("response_body")
                if response_body is not None:
                    if isinstance(response_body, str):
                        print(response_body)
                    else:
                        print(json.dumps(response_body, ensure_ascii=False))

                emit_progress(
                    progress_callback,
                    stage="uploading",
                    current=completed_count,
                    total=total_selected,
                    message=(
                        f"Subidos {completed_count}/{total_selected}: {upload_name}"
                        if item.get("uploaded")
                        else f"Fallo {completed_count}/{total_selected}: {upload_name}"
                    ),
                )
                if item_callback:
                    item_callback("completed", item, completed_count, total_selected)

                _submit_next(executor, pending_iterator)

    uploaded_failed = max(0, len(selected) - uploaded_ok)
    upload_items.sort(key=lambda item: item.get("_order", 0))
    for item in upload_items:
        item.pop("_order", None)
    return {
        "selected": len(selected),
        "uploaded_ok": uploaded_ok,
        "uploaded_failed": uploaded_failed,
        "items": upload_items,
    }


def run_seia_notebook_pipeline(
    documento_seia,
    id_adenda=None,
    tipo=None,
    output_dir=None,
    output_base_dir=DEFAULT_OUTPUT,
    skip_size_estimation=False,
    no_extract=False,
    keep_existing=False,
    enable_download=None,
    upload_limit=NOTEBOOK_UPLOAD_LIMIT,
    notebook_title=None,
    notebook_api_base_url=None,
    require_notebook=False,
    on_notebook_created: Optional[Callable[[str, str], None]] = None,
    stop_after_cp6b=False,
    exclude_keywords=None,
    progress_callback=None,
    notebook_auth=None,
):
    """
    Ejecuta pipeline completo CP1-CP8 y retorna un resumen estructurado.
    """
    url = validate_url(documento_seia)
    id_doc = extract_id_documento(url)
    global_start = time.time()

    # ==== HEADER ====
    print()
    print("=" * 64)
    print("SEIA Document Downloader")
    print("=" * 64)
    print(f"  URL: {url}")
    print(f"  ID Documento: {id_doc}")
    if id_adenda is not None:
        print(f"  ID Adenda: {id_adenda}")
    if tipo:
        print(f"  Tipo: {tipo}")

    session = create_session()
    auth_seed = None
    if notebook_auth is not None:
        auth_seed = prepare_notebook_client_seed(notebook_auth)
    emit_progress(
        progress_callback,
        stage="fetching",
        current=0,
        total=1,
        message="Obteniendo pagina del documento SEIA.",
    )

    # ==== CHECKPOINT 1: Fetch pagina ====
    print()
    print("[CP1] Obteniendo pagina...")
    soup = fetch_page(session, url)

    metadata = extract_metadata(soup)
    print(f"  Proyecto:  {metadata['nombre_proyecto'] or '(no detectado)'}")
    print(f"  Tipo:      {metadata['tipo_documento'] or '(no detectado)'}")
    if metadata["empresa"]:
        print(f"  Empresa:   {metadata['empresa']}")
    if metadata["region"]:
        print(f"  Region:    {metadata['region']}")
    if metadata["inversion"]:
        print(f"  Inversion: {metadata['inversion']}")

    # ==== CHECKPOINT 2: Descubrimiento ====
    print()
    print("[CP2] Buscando documentos...")
    emit_progress(
        progress_callback,
        stage="discovering",
        current=0,
        total=1,
        message="Buscando documentos descargables.",
    )
    documents = extract_download_links(soup, url)

    if not documents:
        raise RuntimeError("No se encontraron documentos descargables en esta pagina.")

    # Estadisticas por tipo
    type_counter = Counter(d["file_type"] for d in documents)
    sections = set(d["section"] for d in documents)

    print(f"  Documentos encontrados: {len(documents)}")
    print(f"  Secciones: {len(sections)}")
    print(f"  Por tipo:")
    for ext, count in type_counter.most_common():
        ext_display = ext if ext else "(sin ext)"
        print(f"    {ext_display:8s}  {count} archivos")

    # Estimacion de tamanos
    if not skip_size_estimation:
        print()
        print("  Estimando tamanos...")
        total_est_size, est_count = estimate_all_sizes(session, documents)
        print()
        print(f"  Tamano total estimado: ~{format_size(total_est_size)} "
              f"({est_count}/{len(documents)} archivos con tamano conocido)")

        # Estadisticas por tipo con tamano
        print(f"  Por tipo (con tamano):")
        for ext, count in type_counter.most_common():
            ext_display = ext if ext else "(sin ext)"
            ext_size = sum(d.get("size_bytes") or 0 for d in documents if d["file_type"] == ext)
            print(f"    {ext_display:8s}  {count} archivos  (~{format_size(ext_size)})")

    # ==== CHECKPOINT 3: Descarga ====
    output_base_dir = Path(output_base_dir)
    if output_dir is None:
        project_name = compact_component(metadata["nombre_proyecto"] or "proyecto", maxlen=42)
        id_doc_dir = compact_component(id_doc or "desconocido", maxlen=28)
        output_dir = output_base_dir / f"{id_doc_dir}_{project_name}"
        base_output_dir = output_base_dir
    else:
        output_dir = Path(output_dir)
        base_output_dir = output_dir.parent
    done, failed, total_bytes, dl_elapsed = 0, 0, 0, 0.0
    download_enabled = ENABLE_DOWNLOAD if enable_download is None else bool(enable_download)
    extraction_levels = {}
    extraction_ran = False
    ext_ok, ext_fail, ext_files = 0, 0, 0

    output_dir = output_dir.resolve()
    existing_manifest_path = output_dir / "manifest.json"
    existing_manifest_data = load_manifest(existing_manifest_path)
    existing_manifest_docs = (
        (existing_manifest_data or {}).get("documentos", [])
        if isinstance(existing_manifest_data, dict)
        else []
    )

    if download_enabled:
        output_dir = prepare_output_directory(
            output_dir=output_dir,
            base_output_dir=base_output_dir,
            clean_start=not keep_existing,
        )

        print()
        if keep_existing:
            print("  Modo salida: conservar archivos existentes.")
        else:
            print("  Modo salida: limpieza inicial activada (sin duplicados de corridas previas).")
        print()
        print(f"[CP3] Descargando {len(documents)} archivos a {output_dir}/")
        done, failed, total_bytes, dl_elapsed = download_all(
            session, documents, output_dir, progress_callback=progress_callback
        )

        # ==== CHECKPOINT 4: Extraccion de archivos comprimidos ====
        if not no_extract:
            print()
            print(f"[CP4] Extrayendo archivos comprimidos...")
            emit_progress(
                progress_callback,
                stage="extracting",
                current=0,
                total=1,
                message="Extrayendo archivos comprimidos.",
            )
            fixed_ext = repair_downloaded_file_extensions(output_dir, documents)
            if fixed_ext:
                print(f"  Correcciones de extension aplicadas: {fixed_ext}")
            ext_ok, ext_fail, ext_files, extraction_levels = extract_all_archives(
                output_dir, documents=documents
            )
            extraction_ran = True
            print()
            print(f"  Archivos extraidos: {ext_ok} OK, {ext_fail} errores, "
                  f"{ext_files} archivos internos")
    else:
        output_dir.mkdir(parents=True, exist_ok=True)
        print()
        print("[CP3] Descarga deshabilitada por configuracion (ENABLE_DOWNLOAD=False).")
        print("  Se omite descarga.")
        if not no_extract:
            print()
            print("[CP4] Extrayendo archivos comprimidos existentes en carpeta...")
            emit_progress(
                progress_callback,
                stage="extracting",
                current=0,
                total=1,
                message="Extrayendo archivos comprimidos existentes.",
            )
            docs_for_repair = [
                d for d in documents
                if d.get("status") == "done" and d.get("download_path")
            ]
            if not docs_for_repair:
                docs_for_repair = [
                    d for d in existing_manifest_docs
                    if d.get("status") == "done" and d.get("download_path")
                ]
            fixed_ext = repair_downloaded_file_extensions(output_dir, docs_for_repair)
            if fixed_ext:
                print(f"  Correcciones de extension aplicadas: {fixed_ext}")
            ext_ok, ext_fail, ext_files, extraction_levels = extract_all_archives(
                output_dir, documents=docs_for_repair
            )
            extraction_ran = True
            print()
            print(f"  Archivos extraidos: {ext_ok} OK, {ext_fail} errores, "
                  f"{ext_files} archivos internos")
        else:
            print("  Se omite extraccion.")

    # ==== CHECKPOINT 5: Resumen + metricas finales ====
    elapsed_before_notebook = time.time() - global_start

    # Guardar/preservar manifest
    if (not download_enabled) and existing_manifest_data:
        manifest_path = existing_manifest_path
        print()
        print("[CP5] Manifest existente preservado (modo sin descarga).")
    else:
        manifest_path = save_manifest(documents, metadata, url, output_dir, elapsed_before_notebook)

    reference_documents = [
        d for d in documents
        if d.get("status") == "done" and d.get("download_path")
    ]
    if not reference_documents and existing_manifest_docs:
        reference_documents = [
            d for d in existing_manifest_docs
            if d.get("status") == "done" and d.get("download_path")
        ]

    # Imprimir resumen de descarga
    print_summary(documents, metadata, output_dir, elapsed_before_notebook)

    # Metricas post-extraccion (incluye archivos extraidos)
    if extraction_ran:
        print_extraction_metrics(output_dir)

    # ==== CHECKPOINT 6: Trazabilidad descarga/descompresion ====
    if not reference_documents:
        print()
        print("[CP6] ADVERTENCIA: no hay documentos base con status=done y download_path.")
        print("      La trazabilidad puede quedar incompleta para nivel 0 en esta corrida.")

    trace_rows, trace_stats = build_download_trace_rows(
        output_dir=output_dir,
        reference_documents=reference_documents,
        extraction_levels=extraction_levels,
        return_stats=True,
    )
    print()
    print("[CP6] Resumen trazabilidad:")
    print(f"  Archivos detectados:       {trace_stats['archivos_detectados']}")
    print(f"  Archivos excluidos:        {trace_stats['archivos_excluidos']}")
    print(f"  Finales en trazabilidad:   {trace_stats['archivos_finales_trazabilidad']}")

    trace_excel_path, trace_excel_error = export_download_trace_excel(trace_rows, output_dir)
    print()
    if trace_excel_error:
        print(f"[CP6] ERROR Trazabilidad Excel: {trace_excel_error}")
    else:
        print(f"[CP6] Excel trazabilidad: {trace_excel_path} ({len(trace_rows)} filas)")

    # ==== CHECKPOINT 6B: Listado final PDF desde trazabilidad + Excel ====
    emit_progress(
        progress_callback,
        stage="tracing",
        current=0,
        total=max(1, len(trace_rows)),
        message="Construyendo trazabilidad y listado final.",
    )
    if exclude_keywords is None:
        docs_report, docs_report_stats = build_final_pdf_report_from_trace(trace_rows)
    else:
        docs_report, docs_report_stats = build_final_pdf_report_from_trace(
            trace_rows, exclude_keywords=exclude_keywords
        )
    tipo_value = clean_text(tipo or "")
    if tipo_value:
        for item in docs_report:
            item["tipo"] = tipo_value
    renamed_final_docs = rename_final_pdf_documents(output_dir, docs_report)
    print_final_pdf_report(docs_report, docs_report_stats)
    if renamed_final_docs:
        print(f"  Renombrados finales aplicados: {renamed_final_docs}")

    excel_path, excel_error = export_final_pdf_report_excel(docs_report, output_dir)
    print()
    if excel_error:
        print(f"[CP6B] ERROR Excel: {excel_error}")
    else:
        print(f"[CP6B] Excel generado: {excel_path}")

    post_metrics = collect_file_metrics(output_dir)
    print_post_filter_summary(
        initial_type_counter=type_counter,
        post_metrics=post_metrics,
        filter_stats=docs_report_stats,
        exclude_keywords=exclude_keywords,
    )

    if stop_after_cp6b:
        total_elapsed = time.time() - global_start
        emit_progress(
            progress_callback,
            stage="persisting_listing",
            current=len(docs_report),
            total=max(1, len(docs_report)),
            message="Guardando listado CP6B.",
            documents_found=len(docs_report),
        )
        print()
        print("[CP7] Se omite creacion de Notebook (stop_after_cp6b=True).")
        print("[CP8] Se omite carga de documentos.")
        print()
        print("Listo.")
        return {
            "status": "listed",
            "id_adenda": id_adenda,
            "tipo": tipo,
            "id_documento": id_doc,
            "notebooklm_id": None,
            "nombre_notebooklm": None,
            "documents_found": len(docs_report),
            "documents_uploaded_ok": 0,
            "documents_uploaded_failed": 0,
            "output_dir": str(Path(output_dir).resolve()),
            "elapsed_seconds": round(total_elapsed, 2),
            "manifest_path": str(Path(manifest_path).resolve()),
            "excel_path": str(Path(excel_path).resolve()) if excel_path else None,
            "trace_excel_path": str(Path(trace_excel_path).resolve()) if trace_excel_path else None,
            "metadata": metadata,
            "notebook_error": None,
            "download_enabled": download_enabled,
            "downloaded_ok": done,
            "downloaded_failed": failed,
            "downloaded_bytes": total_bytes,
            "download_elapsed_seconds": round(dl_elapsed, 2),
            "docs_report": docs_report,
            "docs_report_stats": docs_report_stats,
            "trace_stats": trace_stats,
            "upload_items": [],
        }

    notebook_id = None
    notebook_title_used = None
    notebook_error = None
    upload_stats = {"selected": 0, "uploaded_ok": 0, "uploaded_failed": 0}

    # ==== CHECKPOINT 7/8: Integracion Notebook (opcional) ====
    if ENABLE_NOTEBOOK_SYNC:
        emit_progress(
            progress_callback,
            stage="creating_notebook",
            current=0,
            total=1,
            message="Creando notebook.",
        )
        notebook_id, notebook_title_used, notebook_error = notify_notebook_api(
            notebook_title=notebook_title,
            api_base_url=notebook_api_base_url,
            raise_on_error=require_notebook,
            notebook_auth=notebook_auth,
            auth_seed=auth_seed,
        )

        if notebook_id and on_notebook_created:
            on_notebook_created(notebook_id, notebook_title_used)

        if notebook_id:
            # El listado final CP6B define exactamente los archivos a cargar al notebook.
            upload_candidates = list(docs_report)
            emit_progress(
                progress_callback,
                stage="uploading",
                current=0,
                total=max(1, len(upload_candidates)),
                message="Subiendo documentos al notebook.",
            )
            upload_stats = upload_documents_batch_and_single(
                notebook_id=notebook_id,
                docs_report=upload_candidates,
                limit=upload_limit,
                api_base_url=notebook_api_base_url,
                notebook_auth=notebook_auth,
                auth_seed=auth_seed,
            )
        else:
            print()
            print("[CP8] Se omite carga de documentos porque no se obtuvo notebook_id.")
            if require_notebook:
                raise NotebookAPIError(notebook_error or "No se obtuvo notebook_id.")
    else:
        print()
        print("[CP7] Integracion Notebook deshabilitada por configuracion (ENABLE_NOTEBOOK_SYNC=False).")
        print("[CP8] Se omite carga de documentos.")

    total_elapsed = time.time() - global_start
    pipeline_status = "success" if upload_stats["uploaded_failed"] == 0 else "partial_success"

    print()
    print("Listo.")
    emit_progress(
        progress_callback,
        stage="completed",
        current=upload_stats["uploaded_ok"],
        total=max(1, upload_stats["selected"] or len(docs_report)),
        message="Proceso completado.",
    )

    return {
        "status": pipeline_status,
        "id_adenda": id_adenda,
        "tipo": tipo,
        "id_documento": id_doc,
        "notebooklm_id": notebook_id,
        "nombre_notebooklm": notebook_title_used,
        "documents_found": len(docs_report),
        "documents_uploaded_ok": upload_stats["uploaded_ok"],
        "documents_uploaded_failed": upload_stats["uploaded_failed"],
        "output_dir": str(Path(output_dir).resolve()),
        "elapsed_seconds": round(total_elapsed, 2),
        "manifest_path": str(Path(manifest_path).resolve()),
        "excel_path": str(Path(excel_path).resolve()) if excel_path else None,
        "trace_excel_path": str(Path(trace_excel_path).resolve()) if trace_excel_path else None,
        "metadata": metadata,
        "notebook_error": notebook_error,
        "download_enabled": download_enabled,
        "downloaded_ok": done,
        "downloaded_failed": failed,
        "downloaded_bytes": total_bytes,
        "download_elapsed_seconds": round(dl_elapsed, 2),
        "docs_report": docs_report,
        "docs_report_stats": docs_report_stats,
        "trace_stats": trace_stats,
        "upload_items": upload_stats.get("items", []),
    }


def main():
    args = parse_args()
    try:
        run_seia_notebook_pipeline(
            documento_seia=args.url,
            id_adenda=None,
            output_dir=None,
            output_base_dir=args.output,
            skip_size_estimation=args.skip_size_estimation,
            no_extract=args.no_extract,
            keep_existing=args.keep_existing,
            enable_download=ENABLE_DOWNLOAD,
            upload_limit=NOTEBOOK_UPLOAD_LIMIT,
            notebook_title=None,
            notebook_api_base_url=None,
            require_notebook=False,
        )
    except Exception as e:
        print()
        print(f"ERROR: {e}")
        if "No se encontraron documentos descargables" in str(e):
            sys.exit(0)
        sys.exit(2)


if __name__ == "__main__":
    main()
